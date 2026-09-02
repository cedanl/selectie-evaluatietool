"""
Config wizard: auto-detect configuratie-instellingen uit een selectiedata Excel-bestand.

Bevat drie onderdelen:
1. Detectiefuncties (puur Python, geen Dash)
2. Dash layout (maak_wizard_layout)
3. Callbacks (registreer_callbacks)
"""

import io
import json
import math
import re

import pandas as pd
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from dash import dcc, html, dash_table, Input, Output, State, ctx
import dash
import dash_bootstrap_components as dbc

from transformatie import _decode_upload


# =============================================================================
# 1. Detectiefuncties
# =============================================================================


def detecteer_header_rij(xls: pd.ExcelFile, sheet: str | int = 0) -> int:
    """Scan rijen 0-9; return 1-based index van de eerste 'echte' headerrij."""
    df_raw = xls.parse(sheet_name=sheet, header=None, nrows=10)
    best_row = 0
    best_score = 0
    ncols = min(20, len(df_raw.columns))
    for i in range(len(df_raw)):
        row = df_raw.iloc[i, :ncols]
        filled = sum(1 for v in row if pd.notna(v) and str(v).strip() != "")
        str_cells = sum(
            1 for v in row if pd.notna(v) and isinstance(v, str) and v.strip() != ""
        )
        score = filled + str_cells
        if score > best_score:
            best_score = score
            best_row = i
    return best_row + 1


_ID_PATRONEN = [
    "studentnummer",
    "aanvraagnummer",
    "nummer_aanvraag",
    "kandidaatnummer",
    "deelnemernummer",
    "student_id",
]


def detecteer_id_kolom(headers: list[str]) -> str | None:
    for patroon in _ID_PATRONEN:
        for h in headers:
            if patroon in str(h).lower().replace(" ", ""):
                return h
    return None


def detecteer_totaalscore(headers: list[str]) -> str | None:
    _EXACT = {"totaal", "totaalscore", "total", "totalscore", "total score"}
    candidates = []
    for h in headers:
        lower = str(h).lower().strip()
        if lower in _EXACT:
            return h
        if "totaal" in lower or "total" in lower:
            candidates.append(h)
    if not candidates:
        return None
    suffix_matches = [
        h
        for h in candidates
        if str(h).lower().endswith(("totaal", "total", "totaalscore", "totalscore"))
    ]
    return suffix_matches[-1] if suffix_matches else candidates[-1]


_SKIP_WOORDEN = {
    "data",
    "selectie",
    "selectiedata",
    "dummy",
    "totaalscores",
    "scores",
    "score",
    "ranking",
    "met",
    "formules",
    "en",
    "van",
    "de",
    "het",
    "beoordelingen",
    "master",
    "bachelor",
    "sheet",
    "blad",
    "resultaten",
    "overzicht",
    "export",
    "rapport",
    "tabel",
    "lijst",
    "bestand",
}


def detecteer_metadata(bestandsnaam: str, bladnamen: list[str]) -> dict:
    """Raad opleiding, instelling en jaar op basis van bestandsnaam en bladnamen."""
    tekst = bestandsnaam + " " + " ".join(bladnamen)

    jaren = re.findall(r"20\d{2}", tekst)
    jaar = jaren[0] if jaren else ""

    naam_zonder_ext = re.sub(r"\.\w+$", "", bestandsnaam)
    woorden = re.findall(r"[A-Za-zÀ-ɏ]+", naam_zonder_ext)
    inhoudelijk = [
        w
        for w in woorden
        if w.lower() not in _SKIP_WOORDEN
        and not re.fullmatch(r"20\d{2}", w)
        and len(w) >= 2
    ]

    return {
        "opleiding": " ".join(inhoudelijk) if inhoudelijk else "",
        "instelling": "",
        "jaar": jaar,
    }


_EXCLUDEER_PATRONEN = [
    re.compile(p, re.IGNORECASE)
    for p in [
        r"datum",
        r"tijd",
        r"voltooid",
        r"start",
        r"naam",
        r"email",
        r"e.?mail",
        r"telefoon",
        r"geboort",
        r"adres",
        r"bsn",
        r"rangnummer",
        r"rang.?nummer",
        r"loting",
        r"random",
        r"z.?score",
        r"zscore",
        r"normscore",
        r"norm.?groep",
        r"percentiel",
        r"proctoring",
        r"opmerking",
        r"toelichting",
        r"aanmak",
        r"aanvraag",
        r"afnametaal",
        r"beoordelingsresultaat",
        r"procesvoltooid",
        r"testvoltooid",
        r"aantal.?woord",
    ]
]


def _moet_uitsluiten(kolom_naam: str) -> bool:
    for patroon in _EXCLUDEER_PATRONEN:
        if patroon.search(kolom_naam):
            return True
    return False


def _raad_instrument(kolom: str, alle_kolommen: list[str]) -> str:
    """Raad het instrument op basis van het gedeelde prefix van kolomnamen.

    Zoekt het langste prefix voor het eerste scheidingsteken (_, spatie of -)
    dat bij minimaal twee kolommen voorkomt. Geen hardcoded afkortingen.
    """
    for sep in ("_", " ", "-"):
        if sep in kolom:
            prefix = kolom.split(sep)[0]
            if len(prefix) >= 2:
                count = sum(1 for k in alle_kolommen if k.split(sep)[0] == prefix)
                if count >= 2:
                    return prefix.replace("_", " ").replace("-", " ").strip().title()

    return ""


def _maak_item_naam(kolom: str) -> str:
    """Leid een leesbare itemnaam af uit de kolomnaam.

    Verwijdert het instrumentprefix (alles voor het eerste scheidingsteken)
    als dat prefix bij meerdere kolommen hoort, plus suffixen als 'score'
    en 'schaalscore'. Geen hardcoded afkortingen.
    """
    naam = kolom

    for sep in ("_", " ", "-"):
        if sep in naam:
            naam = naam.split(sep, 1)[1]
            break

    naam = re.sub(r"_?[Ss]chaalscore$", "", naam)
    naam = re.sub(r"_?SCORE$", "", naam)
    naam = re.sub(r"_?[Ss]core$", "", naam)

    naam = naam.replace("_", " ").replace("-", " ").strip()
    naam = re.sub(r"([a-z])([A-Z])", r"\1 \2", naam)

    if naam:
        naam = naam[0].upper() + naam[1:]

    return naam


# Nette bovengrenzen waar de schaalsuggestie naar afrondt. Daarboven naar tientallen.
_NETTE_MAXIMA = [3, 5, 7, 10, 20, 25, 50, 100]


def _rond_schaal_max(hoog: float) -> int:
    """Rond de hoogste waarde omhoog naar een nette bovengrens."""
    for n in _NETTE_MAXIMA:
        if hoog <= n:
            return n
    return math.ceil(hoog / 10) * 10


def _raad_schaal(waarden: pd.Series) -> str:
    """Raad een nette schaal 'min-max' uit de scorewaardes.

    Het ruwe databereik is meestal niet de bedoelde schaal (niemand scoort de
    uitersten), dus rondt de suggestie af: de bovengrens omhoog naar een nette
    waarde (1-7, 0-100, ...), de ondergrens naar 0 of 1. Geeft "" terug als er
    geen bruikbaar bereik is.
    """
    getallen = pd.to_numeric(waarden, errors="coerce").dropna()
    if getallen.empty or getallen.min() == getallen.max():
        return ""

    hoog = _rond_schaal_max(getallen.max())
    laag = 0 if getallen.min() <= 0 or hoog > 10 else 1
    return f"{laag}-{hoog}"


def detecteer_alle_kolommen(
    df: pd.DataFrame,
    id_kolom: str | None,
    totaalscore_kolom: str | None,
) -> list[dict]:
    """Eén rij per kolom in het selectiebestand, met `_meenemen` aan voor de
    kolommen die als score worden herkend (numeriek, geen ID/totaal/uitsluiting).
    Voor die kolommen worden instrument, item en schaal alvast voorgesteld; de
    overige kolommen komen leeg en uitgevinkt in de tabel."""
    skip = {id_kolom, totaalscore_kolom} - {None}
    alle_kolommen = [str(c) for c in df.columns]
    resultaat = []

    for col in df.columns:
        col_str = str(col)
        is_score = (
            col_str not in skip
            and pd.api.types.is_numeric_dtype(df[col])
            and not _moet_uitsluiten(col_str)
        )
        resultaat.append(
            {
                "kolom_naam": col_str,
                "instrument": _raad_instrument(col_str, alle_kolommen)
                if is_score
                else "",
                "item": _maak_item_naam(col_str) if is_score else "",
                "criterium": "",
                "schaal": _raad_schaal(df[col]) if is_score else "",
                "_meenemen": is_score,
            }
        )

    return resultaat


def _instrument_tip(score_kols: list[dict]):
    """Geef een tip als de data instrument-level scores heeft (1 kolom per instrument).

    Dit is herkenbaar doordat het instrument veld leeg is (geen gedeeld prefix),
    of doordat elk instrument slechts 1 item heeft. Alleen meegenomen
    scorekolommen tellen mee.
    """
    score_kols = [r for r in score_kols if r.get("_meenemen", True)]
    instrumenten = {}
    for rij in score_kols:
        inst = rij.get("instrument", "").strip()
        if inst:
            instrumenten.setdefault(inst, []).append(rij["kolom_naam"])

    if not instrumenten:
        geen_instrument = [r for r in score_kols if not r.get("instrument", "").strip()]
        if geen_instrument:
            return dbc.Alert(
                [
                    html.Strong("Tip: "),
                    "De wizard herkent geen instrumentgroepen (gedeelde prefixen). "
                    "Als elke kolom een apart meetinstrument is (bijv. 'Wiskunde', "
                    "'Motivatiebrief', 'Interview'), vul dan bij elke rij het "
                    "instrument in. De itemnaam mag hetzelfde zijn als het instrument.",
                ],
                color="info",
                className="small py-2 mb-0",
            )
        return ""

    alleen_1_item = all(len(items) == 1 for items in instrumenten.values())
    if alleen_1_item and len(instrumenten) >= 2:
        return dbc.Alert(
            [
                html.Strong("Tip: "),
                "Elk instrument heeft precies 1 scorekolom. Dat is prima: de tool "
                "werkt ook met instrumentscores zonder onderliggende items. "
                "Controleer dat de instrument- en itemnamen kloppen.",
            ],
            color="info",
            className="small py-2 mb-0",
        )

    return ""


def bouw_config_dict(
    blad_naam: str,
    header_rij: int,
    koppel_id_kolom: str,
    totaalscore_kolom: str,
    opleiding: str,
    instellingscode: str,
    jaar: str,
    kolommen: list[dict],
) -> dict:
    return {
        "koppel_id_kolom": str(koppel_id_kolom).strip(),
        "opleiding": str(opleiding).strip(),
        "instellingscode": str(instellingscode).strip(),
        "jaar": str(jaar).strip(),
        "blad_naam": str(blad_naam).strip(),
        "header_rij": str(int(header_rij)),
        "totaalscore_kolom": str(totaalscore_kolom).strip(),
        "kolommen": [
            {
                k: (bool(v) if k == "meenemen" else str(v).strip())
                for k, v in kol.items()
            }
            for kol in kolommen
        ],
    }


_INST_UITLEG = {
    "koppel_id_kolom": "Naam van de kolom in je selectiebestand die het studentnummer bevat. Dit is de kolom die gebruikt wordt om selectiedata aan 1CHO-data te koppelen.",
    "opleiding": "Naam of code van de opleiding, bijvoorbeeld 'Psychologie' of 'B PSY'.",
    "instellingscode": "BRIN-code van de instelling, bijvoorbeeld '21RI' voor Radboud.",
    "jaar": "Selectiejaar (het jaar waarin de selectie heeft plaatsgevonden), bijvoorbeeld '2026'.",
    "blad_naam": "Naam van het tabblad in je selectie-Excel dat de scores bevat.",
    "header_rij": "Rijnummer van de kopregel in het selectiebestand (1 = eerste rij). Sommige bestanden hebben meerdere rijen boven de kolomnamen.",
    "totaalscore_kolom": "Optioneel: naam van een kolom die al een totaalscore bevat. Laat leeg als er geen totaalscore-kolom is; de tool berekent dan zelf een totaalscore.",
}

_KOL_UITLEG = {
    "meenemen": "TRUE = deze kolom wordt meegenomen als selectie-item in de analyses. FALSE = kolom wordt genegeerd (bijv. tekstkolommen of id-kolommen).",
    "kolom_naam": "De exacte kolomnaam zoals die in je selectiebestand staat. Niet aanpassen.",
    "instrument": "Het selectie-instrument waaronder dit item valt, bijv. 'Motivatiebrief' of 'Capaciteitentest'. Meerdere items kunnen hetzelfde instrument delen.",
    "item": "De naam van dit specifieke onderdeel, bijv. 'Analytisch vermogen' of 'Totaalscore gesprek'.",
    "criterium": "Optioneel groeperingsniveau tussen instrument en item, bijv. 'Cognitief' of 'Persoonlijkheid'. Laat leeg als je geen extra groepering wilt.",
    "schaal": "Het scorebereik van dit item, bijv. '1-7' of '0-100'. Wordt gebruikt voor de visualisaties.",
}

_HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
_HEADER_FONT = Font(bold=True)
_UITLEG_FONT = Font(color="595959", italic=True)
_UITLEG_ALIGN = Alignment(wrap_text=True)
_KOL_HEADERS = [
    ("meenemen", 12),
    ("kolom_naam", 35),
    ("instrument", 22),
    ("item", 30),
    ("criterium", 18),
    ("schaal", 10),
]


def exporteer_config_excel(config: dict) -> bytes:
    wb = Workbook()

    ws_inst = wb.active
    ws_inst.title = "instellingen"
    ws_inst.column_dimensions["A"].width = 22
    ws_inst.column_dimensions["B"].width = 35
    ws_inst.column_dimensions["C"].width = 80

    inst_rijen = [
        ("koppel_id_kolom", config.get("koppel_id_kolom", "")),
        ("opleiding", config.get("opleiding", "")),
        ("instellingscode", config.get("instellingscode", "")),
        ("jaar", config.get("jaar", "")),
        ("blad_naam", config.get("blad_naam", "")),
        ("header_rij", config.get("header_rij", "1")),
        ("totaalscore_kolom", config.get("totaalscore_kolom", "")),
    ]
    for r, (key, val) in enumerate(inst_rijen, start=1):
        ws_inst.cell(row=r, column=1, value=key).font = _HEADER_FONT
        ws_inst.cell(row=r, column=2, value=val)
        uitleg_cel = ws_inst.cell(row=r, column=3, value=_INST_UITLEG.get(key, ""))
        uitleg_cel.font = _UITLEG_FONT
        uitleg_cel.alignment = _UITLEG_ALIGN

    ws_kol = wb.create_sheet("kolommen")
    for c, (h, w) in enumerate(_KOL_HEADERS, start=1):
        ws_kol.column_dimensions[get_column_letter(c)].width = w
        header_cel = ws_kol.cell(row=1, column=c, value=h)
        header_cel.font = _HEADER_FONT
        header_cel.fill = _HEADER_FILL
        header_cel.comment = Comment(_KOL_UITLEG.get(h, ""), "Config wizard")

    kol_headers = [h for h, _ in _KOL_HEADERS]
    for r, kol in enumerate(config.get("kolommen", []), start=2):
        for c, veld in enumerate(kol_headers, start=1):
            if veld == "meenemen":
                ws_kol.cell(row=r, column=c, value=bool(kol.get("meenemen", True)))
            else:
                ws_kol.cell(row=r, column=c, value=kol.get(veld, ""))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# =============================================================================
# 2. Layout
# =============================================================================


def maak_wizard_layout() -> html.Div:
    return html.Div(
        [
            dcc.Store(id="wiz-config-store", storage_type="memory"),
            dcc.Store(id="wiz-raw-store", storage_type="memory"),
            dcc.Download(id="wiz-download"),
            html.P(
                "Geen configuratiebestand? Laat de wizard er een voor je maken:",
                className="text-muted small mb-1 mt-2",
            ),
            dbc.Button(
                "Config automatisch genereren",
                id="wiz-open-btn",
                color="primary",
                outline=True,
                className="w-100 mb-2",
            ),
            html.Div(
                id="wiz-overlay",
                className="wiz-overlay",
                style={"display": "none"},
                children=html.Div(
                    [
                        html.Div(
                            [
                                html.H5(
                                    "Config automatisch genereren", className="mb-0"
                                ),
                                dbc.Button(
                                    "Sluiten",
                                    id="wiz-close-btn",
                                    color="danger",
                                    size="sm",
                                ),
                            ],
                            className=(
                                "d-flex justify-content-between align-items-center mb-3"
                            ),
                        ),
                        html.P(
                            "Een selectieprocedure is opgebouwd uit een of meer "
                            "meetinstrumenten: bijvoorbeeld een toets, een beoordeling "
                            "van het schooldiploma of een selectiegesprek. Elk "
                            "instrument levert per kandidaat een of meer scores op. De "
                            "tool moet weten hoe jouw selectiebestand die opbouw "
                            "weergeeft; dat leg je vast in een configuratie.",
                            className="wiz-uitleg mb-2",
                        ),
                        html.P(
                            "De wizard leest je selectiebestand en vult dit zoveel "
                            "mogelijk automatisch in. Je controleert het in een paar "
                            "onderdelen: welk werkblad en welke kolomkoprij de tool moet "
                            "gebruiken, welke kolom de kandidaat herkent (om aan de "
                            "1CHO-data te koppelen), en welke kolommen scores zijn. Per "
                            "scorekolom geef je aan bij welk instrument hij hoort, wat "
                            "hij meet (het item), een eventueel criterium en op welke "
                            "schaal hij loopt.",
                            className="wiz-uitleg mb-2",
                        ),
                        html.P(
                            "Loop de velden hieronder na, pas aan waar nodig en klik op "
                            "'Bevestig config'. Daarna upload je de 1CHO-data om het "
                            "dashboard te openen.",
                            className="wiz-uitleg mb-3",
                        ),
                        # Blad en headerrij
                        dbc.Row(
                            [
                                dbc.Col(
                                    [
                                        dbc.Label("Blad", className="small"),
                                        dcc.Dropdown(
                                            id="wiz-sheet-dropdown",
                                            placeholder="Upload eerst selectiedata",
                                            clearable=False,
                                            className="mb-2",
                                        ),
                                    ]
                                ),
                                dbc.Col(
                                    [
                                        dbc.Label("Headerrij", className="small"),
                                        dbc.Input(
                                            id="wiz-header-rij",
                                            type="number",
                                            min=1,
                                            max=20,
                                            value=1,
                                            size="sm",
                                        ),
                                    ],
                                    width=4,
                                ),
                            ],
                            className="mb-2",
                        ),
                        # Opleiding en jaar
                        dbc.Row(
                            [
                                dbc.Col(
                                    [
                                        dbc.Label("Opleiding", className="small"),
                                        dbc.Input(
                                            id="wiz-opleiding",
                                            placeholder="bijv. Farmacie",
                                            size="sm",
                                        ),
                                    ]
                                ),
                                dbc.Col(
                                    [
                                        dbc.Label("Instelling", className="small"),
                                        dbc.Input(
                                            id="wiz-instelling",
                                            placeholder="bijv. LUMC",
                                            size="sm",
                                        ),
                                    ]
                                ),
                                dbc.Col(
                                    [
                                        dbc.Label("Selectiejaar", className="small"),
                                        dbc.Input(
                                            id="wiz-jaar",
                                            placeholder="bijv. 2026",
                                            size="sm",
                                            type="number",
                                        ),
                                    ],
                                    width=3,
                                ),
                            ],
                            className="mb-2",
                        ),
                        # ID-kolom en totaalscore
                        dbc.Row(
                            [
                                dbc.Col(
                                    [
                                        dbc.Label("ID-kolom", className="small"),
                                        dcc.Dropdown(
                                            id="wiz-id-kolom",
                                            placeholder="Wordt automatisch gedetecteerd",
                                            clearable=False,
                                        ),
                                        dbc.FormText(
                                            "De kolom die een student herkent (bijv. "
                                            "studentnummer). Hiermee koppelen we de "
                                            "scores aan de 1CHO-data."
                                        ),
                                    ]
                                ),
                                dbc.Col(
                                    [
                                        dbc.Label(
                                            "Totaalscore-kolom", className="small"
                                        ),
                                        dcc.Dropdown(
                                            id="wiz-totaalscore",
                                            placeholder="Wordt automatisch gedetecteerd",
                                            clearable=True,
                                        ),
                                        dbc.FormText(
                                            "Optioneel: de kolom met de eindscore, "
                                            "als die in je bestand staat."
                                        ),
                                    ]
                                ),
                            ],
                            className="mb-2",
                        ),
                        # Kolommen tabel
                        html.Div(
                            id="wiz-tabel-container",
                            children=[
                                dbc.Label("Kolommen", className="small"),
                                html.P(
                                    [
                                        "Elke rij is een kolom uit je selectiebestand. "
                                        "Het ",
                                        html.Strong("vinkje vooraan elke rij"),
                                        " bepaalt of die kolom meegaat in de analyse. "
                                        "De wizard vinkt de kolommen die hij als score "
                                        "herkent alvast aan; vink een kolom uit die "
                                        "geen selectiescore is, of vink er een aan die "
                                        "de wizard miste. Instrument is het "
                                        "meetinstrument (bijv. een test of "
                                        "beoordeling), Item is wat het meet, "
                                        "Criterium is een optionele groepering, en "
                                        "Schaal is het bereik van de scores (bijv. "
                                        "1-7 of 0-100). Voor herkende scorekolommen "
                                        "stelt de wizard deze velden alvast voor; pas "
                                        "ze aan waar nodig.",
                                    ],
                                    className="wiz-uitleg mb-2",
                                ),
                                html.P(
                                    "Upload selectiedata om kolommen te detecteren.",
                                    id="wiz-tabel-placeholder",
                                    className="small text-muted",
                                ),
                                dash_table.DataTable(
                                    id="wiz-kolommen-tabel",
                                    columns=[
                                        {
                                            "name": "Kolom",
                                            "id": "kolom_naam",
                                            "editable": False,
                                        },
                                        {
                                            "name": "Instrument",
                                            "id": "instrument",
                                            "editable": True,
                                        },
                                        {
                                            "name": "Item",
                                            "id": "item",
                                            "editable": True,
                                        },
                                        {
                                            "name": "Criterium",
                                            "id": "criterium",
                                            "editable": True,
                                        },
                                        {
                                            "name": "Schaal",
                                            "id": "schaal",
                                            "editable": True,
                                        },
                                    ],
                                    data=[],
                                    editable=True,
                                    row_deletable=False,
                                    row_selectable="multi",
                                    selected_rows=[],
                                    style_table={
                                        "overflowX": "auto",
                                        "fontSize": "13px",
                                    },
                                    style_header={
                                        "backgroundColor": "#f8f9fa",
                                        "fontWeight": "600",
                                        "fontSize": "12px",
                                    },
                                    style_cell={
                                        "textAlign": "left",
                                        "padding": "4px 8px",
                                        "whiteSpace": "normal",
                                        "height": "auto",
                                    },
                                    style_data_conditional=[
                                        {
                                            "if": {"column_id": "kolom_naam"},
                                            "backgroundColor": "#f8f9fa",
                                            "color": "#6c757d",
                                        },
                                    ],
                                ),
                            ],
                        ),
                        html.Div(id="wiz-tip", className="mt-2"),
                        html.Div(id="wiz-status", className="mt-2 mb-2"),
                        dbc.Row(
                            [
                                dbc.Col(
                                    dbc.Button(
                                        "Bevestig config",
                                        id="wiz-bevestig-btn",
                                        color="primary",
                                        size="sm",
                                        className="w-100",
                                    ),
                                ),
                                dbc.Col(
                                    dbc.Button(
                                        "Download als Excel",
                                        id="wiz-download-btn",
                                        color="secondary",
                                        size="sm",
                                        outline=True,
                                        className="w-100",
                                        style={"display": "none"},
                                    ),
                                ),
                            ],
                            className="mt-3 g-2",
                        ),
                    ],
                    className="wiz-card",
                ),
            ),
        ],
        className="mb-3",
    )


# =============================================================================
# 3. Callbacks
# =============================================================================


def registreer_callbacks(app: dash.Dash) -> None:

    @app.callback(
        Output("wiz-overlay", "style"),
        Input("wiz-open-btn", "n_clicks"),
        Input("wiz-close-btn", "n_clicks"),
        prevent_initial_call=True,
    )
    def toon_wizard(_open, _close):
        return {"display": "none" if ctx.triggered_id == "wiz-close-btn" else "flex"}

    @app.callback(
        Output("wiz-raw-store", "data"),
        Input("upload-selectiedata", "contents"),
        prevent_initial_call=True,
    )
    def bewaar_selectiedata(contents):
        if not contents:
            return dash.no_update
        return contents

    @app.callback(
        Output("wiz-sheet-dropdown", "options"),
        Output("wiz-sheet-dropdown", "value"),
        Output("wiz-header-rij", "value"),
        Output("wiz-opleiding", "value"),
        Output("wiz-instelling", "value"),
        Output("wiz-jaar", "value"),
        Input("wiz-raw-store", "data"),
        State("upload-selectiedata", "filename"),
        prevent_initial_call=True,
    )
    def detecteer_blad_en_header(raw_contents, filename):
        if not raw_contents:
            return [], None, 1, None, None, None

        try:
            raw = _decode_upload(raw_contents)
            xls = pd.ExcelFile(io.BytesIO(raw))
            bladen = xls.sheet_names
            options = [{"label": b, "value": b} for b in bladen]
            gekozen = bladen[0] if len(bladen) == 1 else None

            header = 1
            if gekozen:
                header = detecteer_header_rij(xls, gekozen)

            meta = detecteer_metadata(filename or "", bladen)

            return (
                options,
                gekozen,
                header,
                meta["opleiding"] or None,
                meta["instelling"] or None,
                meta["jaar"] or None,
            )
        except Exception as e:
            print(f"[wizard] detecteer_blad_en_header mislukt: {e}", flush=True)
            return [], None, 1, None, None, None

    @app.callback(
        Output("wiz-id-kolom", "options"),
        Output("wiz-id-kolom", "value"),
        Output("wiz-totaalscore", "options"),
        Output("wiz-totaalscore", "value"),
        Output("wiz-kolommen-tabel", "data"),
        Output("wiz-kolommen-tabel", "selected_rows"),
        Output("wiz-tabel-placeholder", "style"),
        Output("wiz-tip", "children"),
        Input("wiz-sheet-dropdown", "value"),
        Input("wiz-header-rij", "value"),
        State("wiz-raw-store", "data"),
        prevent_initial_call=True,
    )
    def detecteer_kolommen(blad, header_rij, raw_contents):
        leeg = ([], None, [], None, [], [], {"display": "block"}, "")

        if not blad or not raw_contents or not header_rij:
            return leeg

        try:
            raw = _decode_upload(raw_contents)
            header_idx = int(header_rij) - 1
            df = pd.read_excel(
                io.BytesIO(raw),
                sheet_name=blad,
                header=header_idx,
            )
            headers = [str(c) for c in df.columns]
            col_options = [{"label": h, "value": h} for h in headers]

            id_kol = detecteer_id_kolom(headers)
            totaal_kol = detecteer_totaalscore(headers)

            alle_kols = detecteer_alle_kolommen(df, id_kol, totaal_kol)

            tip = _instrument_tip(alle_kols) if alle_kols else ""

            # De checkboxes (selected_rows) zijn de Meenemen-vlag; vink de
            # herkende scorekolommen vast aan. _meenemen hoort niet in de tabel.
            geselecteerd = [i for i, k in enumerate(alle_kols) if k["_meenemen"]]
            tabel_rijen = [
                {k: v for k, v in kol.items() if k != "_meenemen"} for kol in alle_kols
            ]

            return (
                col_options,
                id_kol,
                col_options,
                totaal_kol,
                tabel_rijen,
                geselecteerd,
                {"display": "none"} if tabel_rijen else {"display": "block"},
                tip,
            )
        except Exception as e:
            print(f"[wizard] detecteer_kolommen mislukt: {e}", flush=True)
            return leeg

    @app.callback(
        Output("wiz-config-store", "data"),
        Output("wiz-status", "children"),
        Output("wiz-download-btn", "style"),
        Input("wiz-bevestig-btn", "n_clicks"),
        State("wiz-kolommen-tabel", "data"),
        State("wiz-kolommen-tabel", "selected_rows"),
        State("wiz-sheet-dropdown", "value"),
        State("wiz-header-rij", "value"),
        State("wiz-id-kolom", "value"),
        State("wiz-totaalscore", "value"),
        State("wiz-opleiding", "value"),
        State("wiz-instelling", "value"),
        State("wiz-jaar", "value"),
        prevent_initial_call=True,
    )
    def bevestig_config(
        n,
        tabel_data,
        selected_rows,
        blad,
        header_rij,
        id_kol,
        totaal_kol,
        opleiding,
        instelling,
        jaar,
    ):
        if not n or not tabel_data:
            return dash.no_update, dash.no_update, {"display": "none"}

        if not id_kol:
            return (
                dash.no_update,
                dbc.Alert(
                    "Selecteer een ID-kolom.", color="danger", className="small py-1"
                ),
                {"display": "none"},
            )

        if not blad:
            return (
                dash.no_update,
                dbc.Alert(
                    "Selecteer een blad.", color="danger", className="small py-1"
                ),
                {"display": "none"},
            )

        # De config bevat alle kolommen; de checkboxes bepalen per rij Meenemen.
        gekozen = set(selected_rows or [])
        kolommen = [
            {**rij, "meenemen": i in gekozen} for i, rij in enumerate(tabel_data)
        ]
        n_actief = sum(k["meenemen"] for k in kolommen)

        if not n_actief:
            return (
                dash.no_update,
                dbc.Alert(
                    "Geen kolommen geselecteerd. Vink minimaal een kolom aan.",
                    color="danger",
                    className="small py-1",
                ),
                {"display": "none"},
            )

        config = bouw_config_dict(
            blad_naam=blad,
            header_rij=header_rij or 1,
            koppel_id_kolom=id_kol,
            totaalscore_kolom=totaal_kol or "",
            opleiding=opleiding or "",
            instellingscode=instelling or "",
            jaar=str(jaar) if jaar else "",
            kolommen=kolommen,
        )

        n_totaal = len(kolommen)
        return (
            json.dumps(config),
            dbc.Alert(
                f"Config aangemaakt ({n_actief} van {n_totaal} kolommen). "
                "Upload nu 1CHO-data om het dashboard te openen.",
                color="success",
                className="small py-1",
            ),
            {"display": "block"},
        )

    @app.callback(
        Output("wiz-download", "data"),
        Input("wiz-download-btn", "n_clicks"),
        State("wiz-config-store", "data"),
        prevent_initial_call=True,
    )
    def download_config(n, config_json):
        if not n or not config_json:
            return dash.no_update
        config = json.loads(config_json)
        return dcc.send_bytes(exporteer_config_excel(config), "config_wizard.xlsx")
