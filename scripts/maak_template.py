import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment

wb = openpyxl.Workbook()

BLUE = "2C3E50"
ACCENT = "2980B9"
LIGHT_BLUE = "D6EAF8"
LIGHT_GRAY = "F2F3F4"
GREEN = "27AE60"
ORANGE = "E67E22"

header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
subheader_font = Font(name="Calibri", size=11, bold=True, color=BLUE)
normal_font = Font(name="Calibri", size=11, color="333333")
example_font = Font(name="Calibri", size=11, color="7F8C8D", italic=True)

header_fill = PatternFill(start_color=BLUE, end_color=BLUE, fill_type="solid")
light_fill = PatternFill(
    start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid"
)
green_fill = PatternFill(start_color="EAFAF1", end_color="EAFAF1", fill_type="solid")

wrap = Alignment(wrap_text=True, vertical="top")
wrap_center = Alignment(wrap_text=True, vertical="center")

thin_border = Border(
    left=Side(style="thin", color="D5D8DC"),
    right=Side(style="thin", color="D5D8DC"),
    top=Side(style="thin", color="D5D8DC"),
    bottom=Side(style="thin", color="D5D8DC"),
)

AUTHOR = "Selectietool"


def style_header_row(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = wrap_center
        cell.border = thin_border


# ============================================================
# TAB 1: UITLEG
# ============================================================
ws_uitleg = wb.active
ws_uitleg.title = "uitleg"
ws_uitleg.sheet_properties.tabColor = ACCENT

ws_uitleg.column_dimensions["A"].width = 30
ws_uitleg.column_dimensions["B"].width = 60
ws_uitleg.column_dimensions["C"].width = 45
ws_uitleg.column_dimensions["D"].width = 30

ws_uitleg.merge_cells("A1:D1")
c = ws_uitleg.cell(row=1, column=1, value="Wat is dit bestand?")
c.font = Font(name="Calibri", size=16, bold=True, color=BLUE)
c.alignment = Alignment(vertical="center")
ws_uitleg.row_dimensions[1].height = 35

ws_uitleg.merge_cells("A2:D2")
c = ws_uitleg.cell(
    row=2,
    column=1,
    value="Dit is een configuratiebestand voor de selectietool. "
    "Je maakt er een per selectiebestand. "
    "Het vertelt de tool hoe jouw Excel-bestand in elkaar zit en welke kolommen meegenomen moeten worden.",
)
c.font = normal_font
c.alignment = wrap
ws_uitleg.row_dimensions[2].height = 40

ws_uitleg.merge_cells("A3:D3")
c = ws_uitleg.cell(
    row=3,
    column=1,
    value='Het configuratiebestand heeft twee tabbladen die je invult: "instellingen" en "kolommen". '
    "In de cellen staan comments met uitleg en voorbeelden. "
    "Klik op een cel met een rood driehoekje rechtsboven om de toelichting te zien. "
    'Dit tabblad ("uitleg") is alleen ter referentie en wordt door de tool genegeerd.',
)
c.font = normal_font
c.alignment = wrap
ws_uitleg.row_dimensions[3].height = 50

# Instellingen uitleg tabel
ws_uitleg.merge_cells("A5:D5")
c = ws_uitleg.cell(row=5, column=1, value="Tabblad: instellingen")
c.font = Font(name="Calibri", size=14, bold=True, color=BLUE)
ws_uitleg.row_dimensions[5].height = 30

for ci, h in enumerate(["Veld", "Wat vul je in?", "Voorbeeld", "Let op"], 1):
    ws_uitleg.cell(row=6, column=ci, value=h)
style_header_row(ws_uitleg, 6, 4)

instellingen_uitleg = [
    [
        "koppel_id_kolom",
        "De exacte kolomnaam die het studentnummer of aanvraagnummer bevat.",
        "Studentnummer\nA_Nummer_Aanvraag",
        "Kopieer exact, inclusief hoofdletters en spaties.",
    ],
    [
        "opleiding",
        "Naam van de opleiding. Vrij in te vullen, wordt label in het dashboard.",
        "Farmacie\nPsychologie",
        "Kies een korte, herkenbare naam.",
    ],
    [
        "jaar",
        "Het selectiejaar.",
        "2025\n2026",
        "Het jaar waarin de selectie plaatsvond.",
    ],
    [
        "blad_naam",
        "De exacte naam van het tabblad (sheet) in het selectiebestand.",
        "2026 LUMC Farmacie\nScores en ranking",
        "Open je bestand en kijk onderin naar de tabbladen.",
    ],
    [
        "header_rij",
        "Rijnummer waarop de kolomnamen staan. Meestal 1.",
        "1\n3",
        "Soms staat er een groepskop boven de kolomnamen, dan is het 2 of 3.",
    ],
    [
        "totaalscore_kolom",
        "De exacte kolomnaam van de definitieve totaalscore.",
        "TOTAALSCORE\nC_Sc_Totaal",
        "Dit is een uitkomst, geen los item.",
    ],
    [
        "rangnummer_kolom",
        "De exacte kolomnaam van het definitieve rangnummer.",
        "Rangnummer definitief\nRangnummer",
        "Het uiteindelijke rangnummer na eventuele loting.",
    ],
    [
        "loting_kolom",
        "Kolomnaam van het lotingsrangnummer. Leeg als niet van toepassing.",
        "Random rangnummer",
        "Niet elk bestand heeft loting.",
    ],
]

for ri, vals in enumerate(instellingen_uitleg, 7):
    for ci, v in enumerate(vals, 1):
        cell = ws_uitleg.cell(row=ri, column=ci, value=v)
        cell.font = normal_font
        cell.alignment = wrap
        cell.border = thin_border
        if ri % 2 == 1:
            cell.fill = light_fill
    ws_uitleg.row_dimensions[ri].height = 45

# Kolommen uitleg tabel
ws_uitleg.merge_cells("A16:D16")
c = ws_uitleg.cell(row=16, column=1, value="Tabblad: kolommen")
c.font = Font(name="Calibri", size=14, bold=True, color=BLUE)
ws_uitleg.row_dimensions[16].height = 30

for ci, h in enumerate(["Veld", "Wat vul je in?", "Voorbeeld", "Let op"], 1):
    ws_uitleg.cell(row=17, column=ci, value=h)
style_header_row(ws_uitleg, 17, 4)

kolommen_uitleg = [
    [
        "kolom_naam",
        "De exacte kolomnaam uit je selectiebestand. Alleen kolommen die hier staan worden meegenomen.",
        "ctb_reflecteren_Schaalscore\nWI SCORE",
        "Kopieer exact uit je bestand.",
    ],
    [
        "instrument",
        "Het meetmiddel. Vrij in te vullen, wordt filter in het dashboard.",
        "Competentietest\nSchooldiploma\nGesprek",
        "Gebruik dezelfde naam voor kolommen van hetzelfde instrument.",
    ],
    [
        "item",
        "Leesbare omschrijving van wat de kolom meet. Dit wordt de naam in het dashboard.",
        "Reflecteren schaalscore\nWiskunde puntenscore",
        "Vermijd technische kolomnamen.",
    ],
    [
        "criterium (optioneel)",
        "De eigenschap die het item meet. Laat leeg als niet van toepassing.",
        "Reflectievermogen\nVakkennis wiskunde",
        "Meerdere items mogen hetzelfde criterium meten.",
    ],
]

for ri, vals in enumerate(kolommen_uitleg, 18):
    for ci, v in enumerate(vals, 1):
        cell = ws_uitleg.cell(row=ri, column=ci, value=v)
        cell.font = normal_font
        cell.alignment = wrap
        cell.border = thin_border
        if ri % 2 == 0:
            cell.fill = light_fill
    ws_uitleg.row_dimensions[ri].height = 55

# Tips
ws_uitleg.merge_cells("A24:D24")
c = ws_uitleg.cell(row=24, column=1, value="Tips")
c.font = Font(name="Calibri", size=14, bold=True, color=BLUE)
ws_uitleg.row_dimensions[24].height = 30

tips = [
    "Neem alleen kolommen op die je wilt analyseren. De rest wordt genegeerd. Liever te weinig dan te veel.",
    "Totaalscore en rangnummer staan in de instellingen, niet in het kolommen-tabblad. Ze zijn uitkomsten, geen losse items.",
    "Als er meerdere versies van een score zijn (bijv. schaalscore en normscore), kies er dan een. Neem niet allebei op.",
    "Als er twee beoordelaars zijn, neem dan de samengevoegde score op, of neem beide apart op (met duidelijke itemnamen).",
    "Tekstvelden, toelichtingen, datums en persoonsgegevens hoef je niet op te nemen.",
    "Test je config: open het selectiebestand en check of elke kolom_naam exact overeenkomt met een kolomnaam in de data.",
]

for i, tip in enumerate(tips, 25):
    ws_uitleg.merge_cells(f"A{i}:D{i}")
    c = ws_uitleg.cell(row=i, column=1, value=f"{i - 24}. {tip}")
    c.font = normal_font
    c.alignment = wrap
    ws_uitleg.row_dimensions[i].height = 30


# ============================================================
# TAB 2: INSTELLINGEN (clean, with comments)
# ============================================================
ws_inst = wb.create_sheet("instellingen")
ws_inst.sheet_properties.tabColor = GREEN

ws_inst.column_dimensions["A"].width = 25
ws_inst.column_dimensions["B"].width = 50

ws_inst.cell(row=1, column=1, value="instelling")
ws_inst.cell(row=1, column=2, value="waarde")
style_header_row(ws_inst, 1, 2)

velden = [
    (
        "koppel_id_kolom",
        "",
        "De exacte kolomnaam die het studentnummer of aanvraagnummer bevat.\n\n"
        "Voorbeelden:\n- Studentnummer\n- A_Nummer_Aanvraag\n\n"
        "Kopieer de kolomnaam exact uit je bestand, inclusief hoofdletters en spaties.",
    ),
    (
        "opleiding",
        "",
        "Naam van de opleiding. Vrij in te vullen.\nWordt gebruikt als label in het dashboard.\n\n"
        "Voorbeelden: Farmacie, Psychologie",
    ),
    ("jaar", "", "Het selectiejaar.\n\nVoorbeelden: 2025, 2026"),
    (
        "blad_naam",
        "",
        "De exacte naam van het tabblad (sheet) in je selectiebestand waar de data staat.\n\n"
        "Open je bestand en kijk onderin welke tabbladen er zijn.\n\n"
        "Voorbeelden:\n- 2026 LUMC Farmacie\n- Scores en ranking\n- 2 Master beoordelingen",
    ),
    (
        "header_rij",
        "1",
        "Het rijnummer waarop de kolomnamen staan.\nMeestal is dit 1.\n\n"
        "Soms staat er een extra rij boven de kolomnamen (bijv. groepskoppen). "
        "Tel dan welke rij de echte kolomnamen heeft.\n\n"
        "Voorbeelden: 1, 3",
    ),
    (
        "totaalscore_kolom",
        "",
        "De exacte kolomnaam van de definitieve totaalscore.\nDit is de score waarop de ranglijst is gebaseerd.\n\n"
        "Dit is een uitkomst van de selectie, geen los item.\n\n"
        "Voorbeelden:\n- TOTAALSCORE\n- C_Sc_Totaal\n- Totale selectiescore %",
    ),
    (
        "rangnummer_kolom",
        "",
        "De exacte kolomnaam van het definitieve rangnummer.\n\n"
        "Voorbeelden:\n- Rangnummer definitief\n- Rangnummer\n- Uiteindelijke rangnummer master FMC 26-27",
    ),
    (
        "loting_kolom",
        "",
        "De kolomnaam van het willekeurig rangnummer bij gelijke scores.\n"
        "Laat LEEG als er geen loting is.\n\n"
        "Voorbeelden:\n- Random rangnummer\n- Random rangnummer bepaald 13-04-2026",
    ),
]

for ri, (veld, default, comment_text) in enumerate(velden, 2):
    cell_a = ws_inst.cell(row=ri, column=1, value=veld)
    cell_a.font = subheader_font
    cell_a.alignment = wrap
    cell_a.border = thin_border

    cell_b = ws_inst.cell(row=ri, column=2, value=default if default else None)
    cell_b.font = normal_font
    cell_b.alignment = wrap
    cell_b.border = thin_border
    cell_b.fill = green_fill
    cell_b.comment = Comment(comment_text, AUTHOR)
    cell_b.comment.width = 350
    cell_b.comment.height = 200

    ws_inst.row_dimensions[ri].height = 28


# ============================================================
# TAB 3: KOLOMMEN (clean, with comments, no example rows)
# ============================================================
ws_kol = wb.create_sheet("kolommen")
ws_kol.sheet_properties.tabColor = ORANGE

ws_kol.column_dimensions["A"].width = 40
ws_kol.column_dimensions["B"].width = 25
ws_kol.column_dimensions["C"].width = 35
ws_kol.column_dimensions["D"].width = 25

headers_info = [
    (
        "kolom_naam",
        "De exacte kolomnaam uit je selectiebestand.\n"
        "De tool zoekt deze kolom op in de data.\n"
        "Alleen kolommen die hier staan worden meegenomen.\n\n"
        "Kopieer exact, inclusief hoofdletters, spaties en regelafbrekingen.\n\n"
        "Voorbeelden:\n- ctb_reflecteren_Schaalscore\n- WI SCORE\n- C_B1_Sc_NL_Docs",
    ),
    (
        "instrument",
        "Het meetmiddel waarmee de score tot stand is gekomen.\n"
        "Vrij in te vullen. Wordt gebruikt als filter in het dashboard.\n\n"
        "Gebruik dezelfde naam voor kolommen die bij hetzelfde instrument horen.\n\n"
        "Voorbeelden:\n- Competentietest\n- Schooldiploma\n- Gesprek\n- Open vraag\n- SIT-S",
    ),
    (
        "item",
        "Een leesbare omschrijving van wat deze kolom concreet meet.\n"
        "Dit wordt de naam in het dashboard.\n\n"
        "Maak het begrijpelijk voor iemand die het bestand niet kent.\n\n"
        "Voorbeelden:\n- Reflecteren schaalscore\n- Wiskunde puntenscore\n- Gespreksbeoordeling (B1)",
    ),
    (
        "criterium (optioneel)",
        "De eigenschap of vaardigheid die dit item meet.\n"
        "Optioneel: laat leeg als het niet van toepassing is (bijv. bij een deelscore).\n\n"
        "Meerdere items mogen hetzelfde criterium meten.\n\n"
        "Voorbeelden:\n- Reflectievermogen\n- Vakkennis wiskunde\n- Communicatievaardigheid\n- Stressbestendigheid",
    ),
]

for ci, (header, comment_text) in enumerate(headers_info, 1):
    cell = ws_kol.cell(row=1, column=ci, value=header)
    cell.comment = Comment(comment_text, AUTHOR)
    cell.comment.width = 350
    cell.comment.height = 250
style_header_row(ws_kol, 1, 4)

for ri in range(2, 52):
    for ci in range(1, 5):
        c = ws_kol.cell(row=ri, column=ci)
        c.border = thin_border
        c.font = normal_font
        c.alignment = wrap

ws_inst.freeze_panes = "A2"
ws_kol.freeze_panes = "A2"
wb.active = 0

wb.save("docs/config_template.xlsx")
print("Template opgeslagen als docs/config_template.xlsx")
