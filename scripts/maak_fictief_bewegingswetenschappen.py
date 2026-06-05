"""
Genereer een fictief selectiebestand voor Bewegingswetenschappen (bachelor).

Structuur gebaseerd op het Psychologie 2026-2027 format: schoolcijfers,
keuzevakken en een matchingsvragenlijst. Simpeler dan het BioMed-bestand,
minder kolommen, ander type selectie.

80 kandidaten, drie instrumenten, geen proceskolommen of proctoring.

Draai:
    uv run python scripts/maak_fictief_bewegingswetenschappen.py
"""

import shutil
import sys
from pathlib import Path

import numpy as np
import pandas as pd

sys.path.insert(0, str(Path(__file__).parent))
from update_configs import make_config

RNG = np.random.default_rng(7777)

OUT_DIR = Path("data/fictief")
OUT_DIR.mkdir(parents=True, exist_ok=True)

N = 80
OPLEIDING = "Bewegingswetenschappen"
INSTELLING = "Vrije Universiteit Amsterdam"
JAAR = 2026
BLAD_NAAM = "Selectiescores 2026"


def clip_round(arr, lo, hi, decimals=2):
    return np.clip(arr, lo, hi).round(decimals)


studentnummers = RNG.choice(range(2000000, 2999999), size=N, replace=False)
studentnummers.sort()

# ── Instrument 1: Schooldiploma (kernvakken + keuzevakken) ───────────────────
# Cijfers op schaal 1-10, puntenscore op schaal 0-5
kernvakken = {
    "Biologie": {"gem": 6.8, "std": 0.9},
    "Scheikunde": {"gem": 6.5, "std": 1.0},
    "Wiskunde A": {"gem": 6.3, "std": 1.1},
}

keuzevakken_pool = [
    "Nederlands", "Engels", "Natuurkunde", "Aardrijkskunde",
    "Economie", "Geschiedenis", "Maatschappijleer",
    "Lichamelijke Opvoeding", "Informatica", "Frans",
]

def cijfer_naar_punten(cijfer):
    return clip_round((cijfer - 4.0) / 1.2, 0, 5)

kolommen = {}
kolom_volgorde = []

for vak, params in kernvakken.items():
    afk = vak[:3].upper()
    cijfers = clip_round(RNG.normal(params["gem"], params["std"], N), 4.0, 10.0)
    punten = cijfer_naar_punten(cijfers)
    kolommen[f"{afk} CIJF"] = cijfers
    kolommen[f"{afk} SCORE"] = punten
    kolom_volgorde.extend([f"{afk} CIJF", f"{afk} SCORE"])

kern_scores = np.column_stack([
    kolommen[f"{vak[:3].upper()} SCORE"] for vak in kernvakken
])
kolommen["BIO+SCH+WIS (0-5)"] = kern_scores.mean(axis=1).round(2)
kolom_volgorde.append("BIO+SCH+WIS (0-5)")

# Keuzevakken: elke student heeft 3-6 keuzevakken
n_keuzevakken_per_student = RNG.integers(3, 7, size=N)
max_keuzevakken = 6

for ki in range(1, max_keuzevakken + 1):
    vak_namen = []
    vak_cijfers = []
    vak_scores = []
    for si in range(N):
        if ki <= n_keuzevakken_per_student[si]:
            vak = RNG.choice(keuzevakken_pool)
            cijf = clip_round(RNG.normal(6.6, 1.0, 1), 4.0, 10.0)[0]
            vak_namen.append(vak)
            vak_cijfers.append(cijf)
            vak_scores.append(cijfer_naar_punten(cijf))
        else:
            vak_namen.append(np.nan)
            vak_cijfers.append(np.nan)
            vak_scores.append(np.nan)
    kolommen[f"K{ki}"] = vak_namen
    kolommen[f"K{ki} CIJF"] = vak_cijfers
    kolommen[f"K{ki} SCORE"] = vak_scores
    kolom_volgorde.extend([f"K{ki}", f"K{ki} CIJF", f"K{ki} SCORE"])

# Combinatiecijfer
alle_cijfers = []
for si in range(N):
    cijs = [kolommen[f"{v[:3].upper()} CIJF"][si] for v in kernvakken]
    for ki in range(1, max_keuzevakken + 1):
        c = kolommen[f"K{ki} CIJF"][si]
        if not np.isnan(c):
            cijs.append(c)
    alle_cijfers.append(np.mean(cijs))

kolommen["Combinatiecijfer"] = np.array(alle_cijfers).round(2)
kolom_volgorde.append("Combinatiecijfer")

# Totaal keuzevakken + combinatiecijfer als deelscore
keuzevak_scores = []
for si in range(N):
    scores = []
    for ki in range(1, max_keuzevakken + 1):
        s = kolommen[f"K{ki} SCORE"][si]
        if not np.isnan(s):
            scores.append(s)
    keuzevak_scores.append(np.mean(scores) if scores else np.nan)

kolommen["Keuzevakken gem (0-5)"] = np.array(keuzevak_scores).round(2)
kolom_volgorde.append("Keuzevakken gem (0-5)")

# ── Instrument 2: Motivatievragenlijst ───────────────────────────────────────
# Score op schaal 1-5
motivatie_score = clip_round(RNG.normal(3.5, 0.8, N), 1.0, 5.0)
kolommen["Motivatie score (1-5)"] = motivatie_score
kolom_volgorde.append("Motivatie score (1-5)")

# ── Instrument 3: Sportvaardigheden assessment ───────────────────────────────
# Praktijkbeoordeling op schaal 1-10
sport_coordinatie = clip_round(RNG.normal(6.8, 1.2, N), 3.0, 10.0)
sport_uithoudingsvermogen = clip_round(RNG.normal(7.0, 1.3, N), 3.0, 10.0)
sport_totaal = ((sport_coordinatie + sport_uithoudingsvermogen) / 2).round(2)

kolommen["Coordinatie (1-10)"] = sport_coordinatie
kolommen["Uithoudingsvermogen (1-10)"] = sport_uithoudingsvermogen
kolommen["Sport totaal (1-10)"] = sport_totaal
kolom_volgorde.extend([
    "Coordinatie (1-10)", "Uithoudingsvermogen (1-10)", "Sport totaal (1-10)"
])

# ── Deelscores en totaalscore ────────────────────────────────────────────────
diploma_pct = (kolommen["BIO+SCH+WIS (0-5)"] / 5 * 100).round(1)
motivatie_pct = (motivatie_score / 5 * 100).round(1)
sport_pct = (sport_totaal / 10 * 100).round(1)

kolommen["Diploma %"] = diploma_pct
kolommen["Motivatie %"] = motivatie_pct
kolommen["Sport %"] = sport_pct
kolom_volgorde.extend(["Diploma %", "Motivatie %", "Sport %"])

totaal_pct = (0.40 * diploma_pct + 0.25 * motivatie_pct + 0.35 * sport_pct).round(1)
kolommen["Totale selectiescore %"] = totaal_pct
kolommen["Rangnummer"] = (
    pd.Series(totaal_pct).rank(ascending=False, method="min").astype(int).values
)
kolom_volgorde.extend(["Totale selectiescore %", "Rangnummer"])

# ── Groepskop-rijen (zoals Psychologie 2026) ─────────────────────────────────
# Het bestand heeft 2 extra rijen boven de kolomnamen: groepskoppen
# header_rij = 3

# ── Build DataFrame ──────────────────────────────────────────────────────────
df = pd.DataFrame({"Studentnummer": studentnummers})
for col in kolom_volgorde:
    df[col] = kolommen[col]

print(f"Selectiebestand: {df.shape[0]} kandidaten, {df.shape[1]} kolommen")

# Write with group headers (like Psychologie 2026)
xlsx_path = OUT_DIR / "selectiedata_bewegingswetenschappen_2026.xlsx"

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

wb = Workbook()
ws = wb.active
ws.title = BLAD_NAAM

header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True)
group_font = Font(bold=True, size=12)

# Row 1: group headers
groups = {
    "A": ("", 1),
    "B": ("Kernvakken Score", 7),
    "H": ("", 1),
    "I": ("Keuzevakken", 18),
}
ws.cell(row=1, column=2, value="Kernvakken Score").font = group_font
ws.cell(row=1, column=9, value="Keuzevakken").font = group_font
ws.cell(row=1, column=28, value="Motivatie").font = group_font
ws.cell(row=1, column=29, value="Sportvaardigheden").font = group_font
ws.cell(row=1, column=32, value="Totale selectiescore en rangnummer").font = group_font

# Row 2: empty (spacer, like the real data)
# Row 3: actual column headers
for ci, col_name in enumerate(df.columns, 1):
    cell = ws.cell(row=3, column=ci, value=col_name)
    cell.font = header_font
    cell.fill = header_fill

# Row 4+: data
for ri, (_, row) in enumerate(df.iterrows(), 4):
    for ci, val in enumerate(row, 1):
        if pd.notna(val):
            ws.cell(row=ri, column=ci, value=val)

wb.save(xlsx_path)
print(f"Opgeslagen: {xlsx_path}")

# ── Configuratiebestand ──────────────────────────────────────────────────────
config_path = OUT_DIR / "config_Bewegingswetenschappen_VU_2026.xlsx"
make_config(
    str(config_path),
    [
        ("Koppel_id_kolom", "Studentnummer"),
        ("opleiding", OPLEIDING),
        ("instellingscode", INSTELLING),
        ("jaar", str(JAAR)),
        ("blad_naam", BLAD_NAAM),
        ("header_rij", "3"),
        ("totaalscore_kolom", "Totale selectiescore %"),
    ],
    [
        ["BIO SCORE", "Schooldiploma", "Biologie puntenscore", "Vakkennis biologie"],
        ["SCH SCORE", "Schooldiploma", "Scheikunde puntenscore", "Vakkennis scheikunde"],
        ["WIS SCORE", "Schooldiploma", "Wiskunde A puntenscore", "Vakkennis wiskunde"],
        ["BIO+SCH+WIS (0-5)", "Schooldiploma", "Gemiddelde kernvakken (0-5)", "Profielsterkheid"],
        ["K1 SCORE", "Schooldiploma keuzevak", "Keuzevak 1 puntenscore", ""],
        ["K2 SCORE", "Schooldiploma keuzevak", "Keuzevak 2 puntenscore", ""],
        ["K3 SCORE", "Schooldiploma keuzevak", "Keuzevak 3 puntenscore", ""],
        ["K4 SCORE", "Schooldiploma keuzevak", "Keuzevak 4 puntenscore", ""],
        ["K5 SCORE", "Schooldiploma keuzevak", "Keuzevak 5 puntenscore", ""],
        ["K6 SCORE", "Schooldiploma keuzevak", "Keuzevak 6 puntenscore", ""],
        ["Keuzevakken gem (0-5)", "Deelscore", "Gemiddelde keuzevakken (0-5)", ""],
        ["Motivatie score (1-5)", "Motivatievragenlijst", "Motivatiescore (1-5)", "Studiemotivatie"],
        ["Coordinatie (1-10)", "Sportvaardigheden", "Coordinatie (1-10)", "Motorische vaardigheden"],
        ["Uithoudingsvermogen (1-10)", "Sportvaardigheden", "Uithoudingsvermogen (1-10)", "Fysieke fitheid"],
        ["Sport totaal (1-10)", "Sportvaardigheden", "Sport totaal (1-10)", ""],
        ["Diploma %", "Deelscore", "Diploma percentage", ""],
        ["Motivatie %", "Deelscore", "Motivatie percentage", ""],
        ["Sport %", "Deelscore", "Sport percentage", ""],
    ],
)

# ── 1CHO-data ────────────────────────────────────────────────────────────────
rang = kolommen["Rangnummer"]
ingeschreven_mask = rang <= 50
ingeschreven_ids = studentnummers[ingeschreven_mask]
ingeschreven_totaal = totaal_pct[ingeschreven_mask]
n_ingeschreven = len(ingeschreven_ids)

totaal_z = (ingeschreven_totaal - ingeschreven_totaal.mean())
if ingeschreven_totaal.std() > 0:
    totaal_z = totaal_z / ingeschreven_totaal.std()
else:
    totaal_z = np.zeros(n_ingeschreven)

doorstroom_kans = 1 / (1 + np.exp(-(0.1 + 0.5 * totaal_z)))
doorstroomt = RNG.random(n_ingeschreven) < doorstroom_kans

geslacht = RNG.choice(["vrouw", "man", "anders"], size=n_ingeschreven, p=[0.55, 0.42, 0.03])
herkomst = RNG.choice(
    ["Nederland", "westerse achtergrond", "Marokko", "Turkije",
     "Suriname/Antillen", "overig niet-westers"],
    size=n_ingeschreven,
    p=[0.74, 0.07, 0.04, 0.03, 0.05, 0.07],
)
vooropleiding = RNG.choice(
    ["VWO", "HAVO + propedeuse", "Anders"],
    size=n_ingeschreven,
    p=[0.78, 0.15, 0.07],
)
vo_cijfers = clip_round(RNG.normal(6.7, 0.55, n_ingeschreven), 5.0, 9.5)

groep = np.where(doorstroomt, "Doorgestroomd naar jaar 2", "Gestart, niet naar jaar 2")

cho_df = pd.DataFrame({
    "studentnummer": ingeschreven_ids,
    "selectiejaar": JAAR,
    "opleiding": OPLEIDING,
    "instellingscode": INSTELLING,
    "groep": groep,
    "geslacht": geslacht,
    "herkomst": herkomst,
    "hoogste_vooropleiding": vooropleiding,
    "gem_eindcijfer_vo": vo_cijfers,
})

cho_path = OUT_DIR / "1cho_data_bewegingswetenschappen_2026.csv"
cho_df.to_csv(cho_path, index=False, sep=";")

print(f"\n1CHO-data: {n_ingeschreven} ingeschreven van {N} kandidaten")
print(f"Groepen: {dict(cho_df['groep'].value_counts())}")

# ── Kopieer naar demo ────────────────────────────────────────────────────────
demo_subdir = Path("data/demo/bewegingswetenschappen_vu_2026")
demo_subdir.mkdir(parents=True, exist_ok=True)
shutil.copy2(xlsx_path, demo_subdir / "selectiedata.xlsx")
shutil.copy2(config_path, demo_subdir / "config.xlsx")
shutil.copy2(cho_path, demo_subdir / "1cho_data.csv")
print(f"Demo bestanden in {demo_subdir}/")

print(f"\nSamenvatting:")
print(f"  Kolommen in selectiebestand: {df.shape[1]}")
print(f"  Kolommen in config: 18")
print(f"  Kandidaten: {N}")
print(f"  header_rij: 3 (groepskoppen boven kolomnamen)")
print(f"  Ingeschreven: {n_ingeschreven}")
print(f"  Doorgestroomd: {int(doorstroomt.sum())}")
print(f"  Niet doorgestroomd: {int(n_ingeschreven - doorstroomt.sum())}")
print(f"  Niet gestart: {N - n_ingeschreven}")
