import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BLUE = "2C3E50"
header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
subheader_font = Font(name="Calibri", size=11, bold=True, color=BLUE)
normal_font = Font(name="Calibri", size=11, color="333333")
header_fill = PatternFill(start_color=BLUE, end_color=BLUE, fill_type="solid")
wrap = Alignment(wrap_text=True, vertical="top")
thin_border = Border(
    left=Side(style="thin", color="D5D8DC"),
    right=Side(style="thin", color="D5D8DC"),
    top=Side(style="thin", color="D5D8DC"),
    bottom=Side(style="thin", color="D5D8DC"),
)


def make_config(filename, instellingen, kolommen):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "instellingen"
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 50

    for ci, h in enumerate(["instelling", "waarde"], 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = header_font
        c.fill = header_fill
        c.border = thin_border

    for ri, (k, v) in enumerate(instellingen, 2):
        ws.cell(row=ri, column=1, value=k).font = subheader_font
        ws.cell(row=ri, column=1).border = thin_border
        c = ws.cell(row=ri, column=2, value=v)
        c.font = normal_font
        c.border = thin_border

    ws2 = wb.create_sheet("kolommen")
    headers = [
        "kolom_naam",
        "instrument",
        "item",
        "criterium (optioneel)",
    ]
    widths = [40, 25, 35, 25]
    for i, w in enumerate(widths):
        ws2.column_dimensions[chr(65 + i)].width = w

    for ci, h in enumerate(headers, 1):
        c = ws2.cell(row=1, column=ci, value=h)
        c.font = header_font
        c.fill = header_fill
        c.border = thin_border

    for ri, row in enumerate(kolommen, 2):
        for ci, v in enumerate(row, 1):
            c = ws2.cell(row=ri, column=ci, value=v)
            c.font = normal_font
            c.alignment = wrap
            c.border = thin_border

    ws.freeze_panes = "A2"
    ws2.freeze_panes = "A2"
    wb.save(filename)
    print(
        "Saved: %s (%d instellingen, %d kolommen)"
        % (filename, len(instellingen), len(kolommen))
    )


# FAR 2025
make_config(
    "config_FAR_Leiden_2025.xlsx",
    [
        ("Koppel_id_kolom", "A_Nummer_Aanvraag"),
        ("opleiding", "Farmacie"),
        ("jaar", "2025"),
        ("blad_naam", "2 Master beoordelingen"),
        ("totaalscore_kolom", "C_Sc_Totaal"),
    ],
    [
        [
            "A_Gem_Cijfer_Bachelor_1",
            "Bachelordiploma",
            "Gemiddeld bachelorcijfer",
            "Studieresultaat",
        ],
        [
            "A_Score_Gem_Cijfer_Ba1",
            "Bachelordiploma",
            "Puntenscore bachelorcijfer",
            "Studieresultaat",
        ],
        [
            "A_Bereken_Pnt_Duur_Ba",
            "Bachelordiploma",
            "Puntenscore studietempo",
            "Studietempo",
        ],
        ["C_B1_Sc_NL_Docs", "Gesprek", "Nederlandse documenten (B1)", "Taalbeheersing"],
        ["C_B2_Sc_NL_Docs", "Gesprek", "Nederlandse documenten (B2)", "Taalbeheersing"],
        [
            "C_B1_Sc_NL_Gespr_Schrijf",
            "Gesprek",
            "Taalvaardigheid gesprek/schrijf (B1)",
            "Taalbeheersing",
        ],
        [
            "C_B2_Sc_NL_Gespr_Schrijf",
            "Gesprek",
            "Taalvaardigheid gesprek/schrijf (B2)",
            "Taalbeheersing",
        ],
        [
            "C_B1_Sc_Beoord_Gesprek",
            "Gesprek",
            "Gespreksbeoordeling (B1)",
            "Communicatievaardigheid",
        ],
        [
            "C_B2_Sc_Beoord_Gesprek",
            "Gesprek",
            "Gespreksbeoordeling (B2)",
            "Communicatievaardigheid",
        ],
        ["C_B1_B2_Sc_SubTotaal", "Gesprek", "Subtotaal gesprek B1+B2", ""],
        ["C_A_Sc_SubTotaal", "Bachelordiploma", "Subtotaal diploma", ""],
    ],
)

# FAR 2026
make_config(
    "config_FAR_Leiden_2026.xlsx",
    [
        ("Koppel_id_kolom", "Studentnummer"),
        ("opleiding", "Farmacie"),
        ("jaar", "2026"),
        ("blad_naam", "2026 LUMC Farmacie"),
        ("totaalscore_kolom", "TOTAALSCORE"),
    ],
    [
        [
            "ctb_reflecteren_Schaalscore",
            "Competentietest",
            "Reflecteren schaalscore",
            "Reflectievermogen",
        ],
        [
            "ctb_schriftelijkcommuniceren_Schaalscore",
            "Competentietest",
            "Schriftelijk communiceren schaalscore",
            "Communicatievaardigheid",
        ],
        [
            "ctb_stressbestendigheid_Schaalscore",
            "Competentietest",
            "Stressbestendigheid schaalscore",
            "Stressbestendigheid",
        ],
        [
            "ctb_initiatiefnemen_Schaalscore",
            "Competentietest",
            "Initiatief nemen schaalscore",
            "Zelfstandigheid",
        ],
        [
            "ctb_doorzettingsvermogen_Schaalscore",
            "Competentietest",
            "Doorzettingsvermogen schaalscore",
            "Zelfstandigheid",
        ],
        [
            "ct_problemenaanpakken_Schaalscore",
            "Competentietest",
            "Problemen aanpakken schaalscore",
            "Copingstijl",
        ],
        [
            "ct_steunzoeken_Schaalscore",
            "Competentietest",
            "Steun zoeken schaalscore",
            "Copingstijl",
        ],
        [
            "ct_problemenvermijden_Schaalscore",
            "Competentietest",
            "Problemen vermijden schaalscore",
            "Copingstijl",
        ],
        [
            "Beoordeling_reflectievermogen\nonder/op/boven niveau = 1/2/3",
            "Open vraag",
            "Beoordeling reflectie (1-2-3)",
            "Reflectievermogen",
        ],
        [
            "Beoordeling_schriftelijkecommunicatie\nonder/op/boven niveau = 1/2/3",
            "Open vraag",
            "Beoordeling communicatie (1-2-3)",
            "Communicatievaardigheid",
        ],
        [
            "sjts_totaal_Schaalscore",
            "SIT-S",
            "Totaalscore sociale intelligentie",
            "Sociale intelligentie",
        ],
        [
            "wetenschappelijkecasusvrijewil_Schaalscore",
            "Wetenschappelijke casus",
            "Schaalscore wetenschappelijke casus",
            "Wetenschappelijk redeneren",
        ],
    ],
)

# Psychologie 2026-2027
make_config(
    "config_Psychologie_2026_2027.xlsx",
    [
        ("Koppel_id_kolom", "Studentnummer"),
        ("opleiding", "Psychologie"),
        ("jaar", "2026"),
        ("blad_naam", "Scores en ranking"),
        ("header_rij", 3),
        ("totaalscore_kolom", "Totale selectiescore %"),
    ],
    [
        ["WI SCORE", "Schooldiploma", "Wiskunde puntenscore", "Vakkennis wiskunde"],
        ["EN SCORE", "Schooldiploma", "Engels puntenscore", "Vakkennis Engels"],
        ["BIO SCORE", "Schooldiploma", "Biologie puntenscore", "Vakkennis biologie"],
        [
            "WI+EN+BIO (0-5)",
            "Schooldiploma",
            "Gemiddelde kernvakken (0-5)",
            "Profielsterkheid",
        ],
        [
            "V1 SCORE",
            "Schooldiploma keuzevak",
            "Keuzevak 1 puntenscore",
            "Vakkennis keuzevak",
        ],
        [
            "V2 SCORE",
            "Schooldiploma keuzevak",
            "Keuzevak 2 puntenscore",
            "Vakkennis keuzevak",
        ],
        [
            "V3 SCORE",
            "Schooldiploma keuzevak",
            "Keuzevak 3 puntenscore",
            "Vakkennis keuzevak",
        ],
        [
            "V4 SCORE",
            "Schooldiploma keuzevak",
            "Keuzevak 4 puntenscore",
            "Vakkennis keuzevak",
        ],
        [
            "V5 SCORE",
            "Schooldiploma keuzevak",
            "Keuzevak 5 puntenscore",
            "Vakkennis keuzevak",
        ],
        [
            "V6 SCORE",
            "Schooldiploma keuzevak",
            "Keuzevak 6 puntenscore",
            "Vakkennis keuzevak",
        ],
        [
            "V7 SCORE",
            "Schooldiploma keuzevak",
            "Keuzevak 7 puntenscore",
            "Vakkennis keuzevak",
        ],
        [
            "V8 SCORE",
            "Schooldiploma keuzevak",
            "Keuzevak 8 puntenscore",
            "Vakkennis keuzevak",
        ],
        [
            "V9 SCORE",
            "Schooldiploma keuzevak",
            "Keuzevak 9 puntenscore",
            "Vakkennis keuzevak",
        ],
        [
            "V10 SCORE",
            "Schooldiploma keuzevak",
            "Keuzevak 10 puntenscore",
            "Vakkennis keuzevak",
        ],
        [
            "COMBICIJF \nSCORE",
            "Schooldiploma",
            "Combinatiecijfer puntenscore",
            "Algemeen studieniveau",
        ],
        ["V1-V10 + COMBICIJF (0-5)", "Deelscore", "Totaalscore vragenlijst (0-5)", ""],
        [
            "Matching Score (1-3)",
            "Matchingsvragenlijst",
            "Matchingscore (1-3)",
            "Studiemotivatie",
        ],
        ["Vragenlijst %", "Deelscore", "Vragenlijst score percentage", ""],
        ["Matching %", "Deelscore", "Matching score percentage", ""],
    ],
)

# NEW: Psychologie 2022-2023
make_config(
    "config_Psychologie_2022_2023.xlsx",
    [
        ("Koppel_id_kolom", "Studentnummer"),
        ("opleiding", "Psychologie"),
        ("jaar", "2022"),
        ("blad_naam", "Totaalscores met formules"),
        ("totaalscore_kolom", "Totaalscore"),
    ],
    [
        ["Toets", "Selectietoets", "Toets ruwe score", "Selectietoets"],
        ["Toetsscore", "Selectietoets", "Toets percentage", "Selectietoets"],
        ["Matching", "Matchingsvragenlijst", "Matching ruwe score", "Studiemotivatie"],
        [
            "Matchingscore",
            "Matchingsvragenlijst",
            "Matching percentage",
            "Studiemotivatie",
        ],
        ["Cijferlijst", "Cijferlijst", "Cijferlijst ruwe score", "Studieniveau"],
        ["Cijferlijstscore", "Cijferlijst", "Cijferlijst percentage", "Studieniveau"],
    ],
)

# Psychologie 2021-2022
make_config(
    "config_Psychologie_2021_2022.xlsx",
    [
        ("Koppel_id_kolom", "Studentnummer"),
        ("opleiding", "Psychologie"),
        ("jaar", "2021"),
        ("blad_naam", "Totaalscores met formules"),
        ("totaalscore_kolom", "Totaalscore"),
    ],
    [
        ["TOETS", "Selectietoets", "Toets ruwe score", "Selectietoets"],
        ["Toets score", "Selectietoets", "Toets percentage", "Selectietoets"],
        ["MATCH", "Matchingsvragenlijst", "Matching ruwe score", "Studiemotivatie"],
        [
            "Match score",
            "Matchingsvragenlijst",
            "Matching percentage",
            "Studiemotivatie",
        ],
        ["CIJF", "Cijferlijst", "Cijferlijst ruwe score", "Studieniveau"],
        ["Cijf score", "Cijferlijst", "Cijferlijst percentage", "Studieniveau"],
    ],
)
