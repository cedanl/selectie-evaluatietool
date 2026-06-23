"""
Genereert drie config.xlsx-bestanden voor BMS (23/24), BMS (24/25, TestVision)
en GEN (22/23). De kolomnamen in deze databronnen zijn heel algemeen, dus
instrument en item zijn ook algemeen gehouden en er zijn geen criteria.

Output: data/configs/config_bms_2324.xlsx, config_bms_2425.xlsx, config_gen_2223.xlsx
"""

from pathlib import Path

from openpyxl import Workbook

UITVOER = Path("data/configs")


def schrijf_config(pad, instellingen, kolommen):
    wb = Workbook()

    ws = wb.active
    ws.title = "instellingen"
    for r, (key, val) in enumerate(instellingen, start=1):
        ws.cell(row=r, column=1, value=key)
        ws.cell(row=r, column=2, value=val)

    ws_kol = wb.create_sheet("kolommen")
    for c, h in enumerate(["kolom_naam", "instrument", "item", "criterium"], start=1):
        ws_kol.cell(row=1, column=c, value=h)
    for r, kol in enumerate(kolommen, start=2):
        ws_kol.cell(row=r, column=1, value=kol[0])
        ws_kol.cell(row=r, column=2, value=kol[1])
        ws_kol.cell(row=r, column=3, value=kol[2])
        ws_kol.cell(row=r, column=4, value="")

    UITVOER.mkdir(parents=True, exist_ok=True)
    wb.save(pad)
    print(f"geschreven: {pad}  ({len(kolommen)} scorekolommen)")


# ---------------------------------------------------------------------------
# 1. BMS 23/24
#    Scorekolommen: Score_1 t/m Score_45, de Big Five (O, C, E, A, N) en
#    Score_Ravens. Result_* en Answer_* zijn geen numerieke scores.
# ---------------------------------------------------------------------------
bms_2324_kol = [
    (f"Score_{i}", "Selectietoets", f"Vraag {i}") for i in range(1, 46)
]
bms_2324_kol += [
    ("O", "Persoonlijkheidsvragenlijst", "O"),
    ("C", "Persoonlijkheidsvragenlijst", "C"),
    ("E", "Persoonlijkheidsvragenlijst", "E"),
    ("A", "Persoonlijkheidsvragenlijst", "A"),
    ("N", "Persoonlijkheidsvragenlijst", "N"),
    ("Score_Ravens", "Capaciteitentest", "Raven"),
]
schrijf_config(
    UITVOER / "config_bms_2324.xlsx",
    [
        ("koppel_id_kolom", "ID"),
        ("opleiding", "BMS"),
        ("instellingscode", ""),
        ("jaar", "2023"),
        ("blad_naam", ""),
        ("header_rij", "1"),
        ("totaalscore_kolom", ""),
    ],
    bms_2324_kol,
)

# ---------------------------------------------------------------------------
# 2. BMS 24/25 (rechtstreeks uit TestVision)
#    Scorekolommen: de genummerde vraagkolommen 1 t/m 30. Toetsscore is de
#    totaalscore. Maxscore/Kansscore/Cesuurscore/Cijfer zijn geen items.
# ---------------------------------------------------------------------------
bms_2425_kol = [(str(i), "Toets", f"Vraag {i}") for i in range(1, 31)]
schrijf_config(
    UITVOER / "config_bms_2425.xlsx",
    [
        ("koppel_id_kolom", "KandidaatID"),
        ("opleiding", "BMS"),
        ("instellingscode", ""),
        ("jaar", "2024"),
        ("blad_naam", ""),
        ("header_rij", "1"),
        ("totaalscore_kolom", "Toetsscore"),
    ],
    bms_2425_kol,
)

# ---------------------------------------------------------------------------
# 3. GEN 22/23
#    Scorekolommen: de ruwe deelscores Deel 1_V1..V8 en Deel 2_V1..V8.
#    De ZDeel*-kolommen zijn al gestandaardiseerd; de tool z-scoort zelf, dus
#    we gebruiken de ruwe kolommen. Totaalscore is de totaalscore.
# ---------------------------------------------------------------------------
gen_2223_kol = [
    (f"Deel 1_V{i}", "Deel 1", f"Vraag {i}") for i in range(1, 9)
]
gen_2223_kol += [
    (f"Deel 2_V{i}", "Deel 2", f"Vraag {i}") for i in range(1, 9)
]
schrijf_config(
    UITVOER / "config_gen_2223.xlsx",
    [
        ("koppel_id_kolom", "code"),
        ("opleiding", "GEN"),
        ("instellingscode", ""),
        ("jaar", "2022"),
        ("blad_naam", ""),
        ("header_rij", "1"),
        ("totaalscore_kolom", "Totaalscore"),
    ],
    gen_2223_kol,
)
