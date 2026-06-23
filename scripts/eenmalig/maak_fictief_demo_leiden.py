"""
Genereer een fictief selectiebestand voor de master Farmacie (demo Leiden).

Structuur lijkt op 'dummy data selectie FAR Leiden 2025': een
masterselectie met een enkele kop-rij (header_rij=1) op het blad
'2 Master beoordelingen'. De selectie bestaat uit een beoordeling van het
bachelordiploma (gemiddeld cijfer en studietempo) en twee onafhankelijke
beoordelaars (B1 en B2) die de Nederlandse documenten, een gesprek/
schrijfopdracht en het selectiegesprek scoren. De puntenscores (C_..._Sc_...)
tellen op tot subtotalen en een eindtotaal (C_Sc_Totaal).

Omdat het een master is, is succes het halen van het diploma in het
cohortjaar (geen doorstroom naar jaar 2). De uitkomstgroep zit niet als
kolom in de data maar wordt door de tool afgeleid uit het 1CHO-bestand.

140 kandidaten, waarvan 60 ingeschreven. Niet herleidbaar naar echte data.

Draai:
    uv run python scripts/eenmalig/maak_fictief_demo_leiden.py
"""

import shutil
import sys
from pathlib import Path

import numpy as np
import pandas as pd

sys.path.insert(0, str(Path(__file__).parent))
from update_configs import make_config

sys.path.insert(0, str(Path(__file__).resolve().parents[2]))
from cho_transform import bouw_ruwe_cho  # noqa: E402

RNG = np.random.default_rng(2468)

OUT_DIR = Path("data/fictief")
OUT_DIR.mkdir(parents=True, exist_ok=True)

N = 140
N_INGESCHREVEN = 60
OPLEIDING = "Farmacie"
INSTELLING = "Universiteit Leiden"
JAAR = 2026
BLAD_NAAM = "2 Master beoordelingen"

aanvraagnummers = RNG.choice(range(200000, 999999), size=N, replace=False)
aanvraagnummers.sort()

# Kolommen worden in volgorde toegevoegd; een insertion-ordered dict bewaart die
# volgorde voor het uiteindelijke DataFrame.
kolommen = {}

# -- Identificatie en achtergrond --------------------------------------------
kolommen["A_Nummer_Aanvraag"] = aanvraagnummers
kolommen["Selectie"] = ["Master"] * N
kolommen["A_Geslacht"] = RNG.choice(["V", "M"], size=N, p=[0.6, 0.4])

bachelor_pool = [
    "Bachelor (university)  -  Farmacie",
    "Bachelor (university)  -  Bio-farmaceutische wetenschappen",
    "Bachelor (university)  -  Scheikunde",
    "Bachelor (university)  -  Biologie",
]
kolommen["A_Gevolgde_Opleiding_1"] = RNG.choice(
    bachelor_pool, size=N, p=[0.55, 0.25, 0.12, 0.08]
)

# -- Bachelordiploma: gemiddeld cijfer en studietempo ------------------------
# Een hoger bachelorcijfer levert een puntenscore (drempel bij een 7).
bachelorcijfer = np.clip(RNG.normal(7.0, 0.5, N), 6.0, 8.5).round(1)
kolommen["A_Gem_Cijfer_Bachelor_1"] = bachelorcijfer
score_cijfer = (bachelorcijfer >= 7.0).astype(int)
kolommen["A_Score_Gem_Cijfer_Ba1"] = score_cijfer

# Studieduur in maanden; 36 is nominaal. Sneller dan nominaal levert 2 punten,
# anders 1 punt.
maanden = np.clip(RNG.normal(38, 4, N), 34, 50).round(0).astype(int)
kolommen["A_Bachelor_Maanden_1"] = maanden
kolommen["A_Mnd_Nominaal1"] = [36] * N
kolommen["A_Mnd_Ba_boven_Nominaal1"] = maanden - 36
score_duur = np.where(maanden <= 38, 2, 1)
kolommen["A_Score_Duur_Ba_Nominaal1"] = score_duur
kolommen["A_Bereken_Pnt_Duur_Ba"] = score_duur


# -- Twee beoordelaars (B1 en B2) --------------------------------------------
def maak_beoordelaar(prefix, achternamen):
    """Genereer de kolommen van een beoordelaar en geef de drie puntenscores
    terug. Elke beoordelaar scoort de Nederlandse documenten (G/V), een gesprek/
    schrijfopdracht (G/V) en het selectiegesprek (2-6). G geeft 2 punten, V 1."""
    kolommen[f"{prefix}_Achternaam"] = RNG.choice(achternamen, size=N)

    nl_docs = RNG.choice(["G", "V"], size=N, p=[0.7, 0.3])
    sc_nl_docs = np.where(nl_docs == "G", 2, 1)
    kolommen[f"{prefix}_NL_Docs"] = nl_docs
    kolommen[f"{prefix}_Score_NL_Docs"] = sc_nl_docs

    nl_gespr = RNG.choice(["G", "V"], size=N, p=[0.65, 0.35])
    sc_nl_gespr = np.where(nl_gespr == "G", 2, 1)
    kolommen[f"{prefix}_NL_Gespr_Schrijf"] = nl_gespr
    kolommen[f"{prefix}_Score_NL_Gespr_Schrijf"] = sc_nl_gespr

    gesprek = RNG.integers(2, 7, size=N)
    kolommen[f"{prefix}_Beoord_Gespr_Schrijfop"] = gesprek
    kolommen[f"{prefix}_Score_Beoord_Gesprek"] = gesprek
    return sc_nl_docs, sc_nl_gespr, gesprek


b1_docs, b1_gespr, b1_gesprek = maak_beoordelaar("B1", ["ZZ", "YY", "XX", "WW", "VV"])
b2_docs, b2_gespr, b2_gesprek = maak_beoordelaar("B2", ["AA", "BB", "CC", "DD", "EE"])

# -- Puntenscores (C-kolommen) en subtotalen ---------------------------------
kolommen["C_Sc_Gem_Cijfer_Ba"] = score_cijfer
kolommen["C_Sc_Duur_Ba_Nominaal1"] = score_duur

kolommen["C_B1_Sc_NL_Docs"] = b1_docs
kolommen["C_B2_Sc_NL_Docs"] = b2_docs
kolommen["C_B1_Sc_NL_Gespr_Schrijf"] = b1_gespr
kolommen["C_B2_Sc_NL_Gespr_Schrijf"] = b2_gespr
kolommen["C_B1_Sc_Beoord_Gesprek"] = b1_gesprek
kolommen["C_B2_Sc_Beoord_Gesprek"] = b2_gesprek

c_b1_subtotaal = b1_docs + b1_gespr + b1_gesprek
c_b2_subtotaal = b2_docs + b2_gespr + b2_gesprek
kolommen["C_B1_Sc_SubTotaal"] = c_b1_subtotaal
kolommen["C_B2_Sc_SubTotaal"] = c_b2_subtotaal
kolommen["C_B1_B2_Sc_SubTotaal"] = c_b1_subtotaal + c_b2_subtotaal

c_a_subtotaal = score_cijfer + score_duur
kolommen["C_A_Sc_SubTotaal"] = c_a_subtotaal

totaal = c_b1_subtotaal + c_b2_subtotaal + c_a_subtotaal
kolommen["C_Sc_Totaal"] = totaal

kolommen["Rangnummer definitief"] = (
    pd.Series(totaal).rank(ascending=False, method="min").astype(int).values
)

# -- Build DataFrame ----------------------------------------------------------
df = pd.DataFrame(kolommen)
print(f"Selectiebestand: {df.shape[0]} kandidaten, {df.shape[1]} kolommen")

# -- Write Excel (header_rij=1, enkele kop-rij) ------------------------------
from openpyxl import Workbook  # noqa: E402
from openpyxl.styles import Font, PatternFill  # noqa: E402

xlsx_path = OUT_DIR / "selectiedata_demo_leiden_2026.xlsx"

wb = Workbook()
ws = wb.active
ws.title = BLAD_NAAM

header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True)

for ci, col_name in enumerate(df.columns, 1):
    cell = ws.cell(row=1, column=ci, value=col_name)
    cell.font = header_font
    cell.fill = header_fill

for ri, (_, row) in enumerate(df.iterrows(), 2):
    for ci, val in enumerate(row, 1):
        if pd.notna(val):
            ws.cell(row=ri, column=ci, value=val)

wb.save(xlsx_path)
print(f"Opgeslagen: {xlsx_path}")

# -- Configuratiebestand ------------------------------------------------------
# Spiegelt de FAR Leiden 2025 config: instrumenten Bachelordiploma en Gesprek,
# met C_Sc_Totaal als totaalscore.
config_path = OUT_DIR / "config_demo_leiden_2026.xlsx"
make_config(
    str(config_path),
    [
        ("Koppel_id_kolom", "A_Nummer_Aanvraag"),
        ("opleiding", OPLEIDING),
        ("instellingscode", INSTELLING),
        ("jaar", str(JAAR)),
        ("blad_naam", BLAD_NAAM),
        ("header_rij", "1"),
        ("totaalscore_kolom", "C_Sc_Totaal"),
    ],
    [
        [
            "A_Gem_Cijfer_Bachelor_1",
            "Bachelordiploma",
            "Gemiddeld bachelorcijfer",
            "Studieresultaat",
        ],
        [
            "A_Score_Gem_Cijfer_Ba1",
            "Bachelordiploma",
            "Puntenscore bachelorcijfer",
            "Studieresultaat",
        ],
        [
            "A_Bereken_Pnt_Duur_Ba",
            "Bachelordiploma",
            "Puntenscore studietempo",
            "Studietempo",
        ],
        ["C_B1_Sc_NL_Docs", "Gesprek", "Nederlandse documenten (B1)", "Taalbeheersing"],
        ["C_B2_Sc_NL_Docs", "Gesprek", "Nederlandse documenten (B2)", "Taalbeheersing"],
        [
            "C_B1_Sc_NL_Gespr_Schrijf",
            "Gesprek",
            "Taalvaardigheid gesprek/schrijf (B1)",
            "Taalbeheersing",
        ],
        [
            "C_B2_Sc_NL_Gespr_Schrijf",
            "Gesprek",
            "Taalvaardigheid gesprek/schrijf (B2)",
            "Taalbeheersing",
        ],
        [
            "C_B1_Sc_Beoord_Gesprek",
            "Gesprek",
            "Gespreksbeoordeling (B1)",
            "Communicatievaardigheid",
        ],
        [
            "C_B2_Sc_Beoord_Gesprek",
            "Gesprek",
            "Gespreksbeoordeling (B2)",
            "Communicatievaardigheid",
        ],
        ["C_B1_B2_Sc_SubTotaal", "Gesprek", "Subtotaal gesprek B1+B2", ""],
        ["C_A_Sc_SubTotaal", "Bachelordiploma", "Subtotaal diploma", ""],
    ],
)

# -- 1CHO-data ----------------------------------------------------------------
# Master: succes is het diploma in het cohortjaar, niet doorstroom naar jaar 2.
rang = kolommen["Rangnummer definitief"]
ingeschreven_mask = rang <= N_INGESCHREVEN
ingeschreven_ids = aanvraagnummers[ingeschreven_mask]
ingeschreven_totaal = totaal[ingeschreven_mask]
n_ingeschreven = len(ingeschreven_ids)

totaal_z = ingeschreven_totaal - ingeschreven_totaal.mean()
if ingeschreven_totaal.std() > 0:
    totaal_z = totaal_z / ingeschreven_totaal.std()
else:
    totaal_z = np.zeros(n_ingeschreven)

diploma_kans = 1 / (1 + np.exp(-(0.4 + 0.6 * totaal_z)))
diploma_behaald = RNG.random(n_ingeschreven) < diploma_kans

geslacht = RNG.choice(["vrouw", "man"], size=n_ingeschreven, p=[0.6, 0.4])
# Bij een master is de vooropleiding een bachelor: vrijwel iedereen hoger
# onderwijs, een enkeling met een buitenlands diploma.
vooropleiding = RNG.choice(
    ["WO Bachelor", "Buitenlands diploma"],
    size=n_ingeschreven,
    p=[0.85, 0.15],
)

cho_df = bouw_ruwe_cho(
    ingeschreven_ids,
    jaar=JAAR,
    diploma_behaald=diploma_behaald,
    opleiding=OPLEIDING,
    instellingscode=INSTELLING,
    geslacht=geslacht,
    vooropleiding_omschrijving=vooropleiding,
)

cho_path = OUT_DIR / "1cho_data_demo_leiden_2026.csv"
cho_df.to_csv(cho_path, index=False, sep=";")

print(f"\n1CHO-data: {n_ingeschreven} ingeschreven van {N} kandidaten")
print(
    f"Diploma gehaald: {int(diploma_behaald.sum())}, "
    f"niet: {int((~diploma_behaald).sum())}"
)

# -- Kopieer naar demo --------------------------------------------------------
demo_subdir = Path("data/demo/demo_leiden_2026")
demo_subdir.mkdir(parents=True, exist_ok=True)
shutil.copy2(xlsx_path, demo_subdir / "selectiedata.xlsx")
shutil.copy2(config_path, demo_subdir / "config.xlsx")
shutil.copy2(cho_path, demo_subdir / "1cho_data.csv")
print(f"Demo bestanden in {demo_subdir}/")

print("\nSamenvatting:")
print(f"  Kolommen in selectiebestand: {df.shape[1]}")
print("  Items in config: 11 (subtotalen meegenomen, ruwe G/V-kolommen niet)")
print(f"  Kandidaten: {N}")
print("  header_rij: 1 (enkele kop-rij)")
print(f"  Ingeschreven: {n_ingeschreven}")
print(f"  Diploma gehaald: {int(diploma_behaald.sum())}")
print(f"  Geen diploma: {int(n_ingeschreven - diploma_behaald.sum())}")
print(f"  Niet gestart: {N - n_ingeschreven}")
