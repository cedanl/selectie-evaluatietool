import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BLUE = "2C3E50"
header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
subheader_font = Font(name="Calibri", size=11, bold=True, color=BLUE)
normal_font = Font(name="Calibri", size=11, color="333333")
header_fill = PatternFill(start_color=BLUE, end_color=BLUE, fill_type="solid")
wrap = Alignment(wrap_text=True, vertical="top")
thin_border = Border(
    left=Side(style="thin", color="D5D8DC"),
    right=Side(style="thin", color="D5D8DC"),
    top=Side(style="thin", color="D5D8DC"),
    bottom=Side(style="thin", color="D5D8DC"),
)


def make_config(filename, instellingen, kolommen):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "instellingen"
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 50

    for ci, h in enumerate(["instelling", "waarde"], 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = header_font
        c.fill = header_fill
        c.border = thin_border

    for ri, (k, v) in enumerate(instellingen, 2):
        ws.cell(row=ri, column=1, value=k).font = subheader_font
        ws.cell(row=ri, column=1).border = thin_border
        c = ws.cell(row=ri, column=2, value=v)
        c.font = normal_font
        c.border = thin_border

    ws2 = wb.create_sheet("kolommen")
    headers = [
        "kolom_naam",
        "instrument",
        "item",
        "criterium (optioneel)",
        "score_type",
    ]
    widths = [40, 25, 35, 25, 20]
    for i, w in enumerate(widths):
        ws2.column_dimensions[chr(65 + i)].width = w

    for ci, h in enumerate(headers, 1):
        c = ws2.cell(row=1, column=ci, value=h)
        c.font = header_font
        c.fill = header_fill
        c.border = thin_border

    for ri, row in enumerate(kolommen, 2):
        for ci, v in enumerate(row, 1):
            c = ws2.cell(row=ri, column=ci, value=v)
            c.font = normal_font
            c.alignment = wrap
            c.border = thin_border

    ws.freeze_panes = "A2"
    ws2.freeze_panes = "A2"
    wb.save(filename)
    print(
        "Saved: %s (%d instellingen, %d kolommen)"
        % (filename, len(instellingen), len(kolommen))
    )


# FAR 2025
make_config(
    "config_FAR_Leiden_2025.xlsx",
    [
        ("Koppel_id_kolom", "A_Nummer_Aanvraag"),
        ("opleiding", "Farmacie"),
        ("jaar", "2025"),
        ("blad_naam", "2 Master beoordelingen"),
        ("totaalscore_kolom", "C_Sc_Totaal"),
    ],
    [
        [
            "A_Gem_Cijfer_Bachelor_1",
            "Bachelordiploma",
            "Gemiddeld bachelorcijfer",
            "Studieresultaat",
            "cijfer (1-10)",
        ],
        [
            "A_Score_Gem_Cijfer_Ba1",
            "Bachelordiploma",
            "Puntenscore bachelorcijfer",
            "Studieresultaat",
            "punten",
        ],
        [
            "A_Bereken_Pnt_Duur_Ba",
            "Bachelordiploma",
            "Puntenscore studietempo",
            "Studietempo",
            "punten",
        ],
        [
            "C_B1_Sc_NL_Docs",
            "Gesprek",
            "Nederlandse documenten (B1)",
            "Taalbeheersing",
            "punten",
        ],
        [
            "C_B2_Sc_NL_Docs",
            "Gesprek",
            "Nederlandse documenten (B2)",
            "Taalbeheersing",
            "punten",
        ],
        [
            "C_B1_Sc_NL_Gespr_Schrijf",
            "Gesprek",
            "Taalvaardigheid gesprek/schrijf (B1)",
            "Taalbeheersing",
            "punten",
        ],
        [
            "C_B2_Sc_NL_Gespr_Schrijf",
            "Gesprek",
            "Taalvaardigheid gesprek/schrijf (B2)",
            "Taalbeheersing",
            "punten",
        ],
        [
            "C_B1_Sc_Beoord_Gesprek",
            "Gesprek",
            "Gespreksbeoordeling (B1)",
            "Communicatievaardigheid",
            "punten",
        ],
        [
            "C_B2_Sc_Beoord_Gesprek",
            "Gesprek",
            "Gespreksbeoordeling (B2)",
            "Communicatievaardigheid",
            "punten",
        ],
        ["C_B1_B2_Sc_SubTotaal", "Gesprek", "Subtotaal gesprek B1+B2", "", "punten"],
        ["C_A_Sc_SubTotaal", "Bachelordiploma", "Subtotaal diploma", "", "punten"],
    ],
)

# FAR 2026
make_config(
    "config_FAR_Leiden_2026.xlsx",
    [
        ("Koppel_id_kolom", "Studentnummer"),
        ("opleiding", "Farmacie"),
        ("jaar", "2026"),
        ("blad_naam", "2026 LUMC Farmacie"),
        ("totaalscore_kolom", "TOTAALSCORE"),
    ],
    [
        [
            "ctb_reflecteren_Schaalscore",
            "Competentietest",
            "Reflecteren schaalscore",
            "Reflectievermogen",
            "schaalscore",
        ],
        [
            "ctb_schriftelijkcommuniceren_Schaalscore",
            "Competentietest",
            "Schriftelijk communiceren schaalscore",
            "Communicatievaardigheid",
            "schaalscore",
        ],
        [
            "ctb_stressbestendigheid_Schaalscore",
            "Competentietest",
            "Stressbestendigheid schaalscore",
            "Stressbestendigheid",
            "schaalscore",
        ],
        [
            "ctb_initiatiefnemen_Schaalscore",
            "Competentietest",
            "Initiatief nemen schaalscore",
            "Zelfstandigheid",
            "schaalscore",
        ],
        [
            "ctb_doorzettingsvermogen_Schaalscore",
            "Competentietest",
            "Doorzettingsvermogen schaalscore",
            "Zelfstandigheid",
            "schaalscore",
        ],
        [
            "ct_problemenaanpakken_Schaalscore",
            "Competentietest",
            "Problemen aanpakken schaalscore",
            "Copingstijl",
            "schaalscore",
        ],
        [
            "ct_steunzoeken_Schaalscore",
            "Competentietest",
            "Steun zoeken schaalscore",
            "Copingstijl",
            "schaalscore",
        ],
        [
            "ct_problemenvermijden_Schaalscore",
            "Competentietest",
            "Problemen vermijden schaalscore",
            "Copingstijl",
            "schaalscore",
        ],
        [
            "Beoordeling_reflectievermogen\nonder/op/boven niveau = 1/2/3",
            "Open vraag",
            "Beoordeling reflectie (1-2-3)",
            "Reflectievermogen",
            "schaal 1-3",
        ],
        [
            "Beoordeling_schriftelijkecommunicatie\nonder/op/boven niveau = 1/2/3",
            "Open vraag",
            "Beoordeling communicatie (1-2-3)",
            "Communicatievaardigheid",
            "schaal 1-3",
        ],
        [
            "sjts_totaal_Schaalscore",
            "SIT-S",
            "Totaalscore sociale intelligentie",
            "Sociale intelligentie",
            "schaalscore",
        ],
        [
            "wetenschappelijkecasusvrijewil_Schaalscore",
            "Wetenschappelijke casus",
            "Schaalscore wetenschappelijke casus",
            "Wetenschappelijk redeneren",
            "schaalscore",
        ],
    ],
)

# Psychologie 2026-2027
make_config(
    "config_Psychologie_2026_2027.xlsx",
    [
        ("Koppel_id_kolom", "Studentnummer"),
        ("opleiding", "Psychologie"),
        ("jaar", "2026"),
        ("blad_naam", "Scores en ranking"),
        ("header_rij", 3),
        ("totaalscore_kolom", "Totale selectiescore %"),
    ],
    [
        [
            "WI SCORE",
            "Schooldiploma",
            "Wiskunde puntenscore",
            "Vakkennis wiskunde",
            "punten (0-5)",
        ],
        [
            "EN SCORE",
            "Schooldiploma",
            "Engels puntenscore",
            "Vakkennis Engels",
            "punten (0-5)",
        ],
        [
            "BIO SCORE",
            "Schooldiploma",
            "Biologie puntenscore",
            "Vakkennis biologie",
            "punten (0-5)",
        ],
        [
            "WI+EN+BIO (0-5)",
            "Schooldiploma",
            "Gemiddelde kernvakken (0-5)",
            "Profielsterkheid",
            "punten (0-5)",
        ],
        [
            "V1 SCORE",
            "Schooldiploma keuzevak",
            "Keuzevak 1 puntenscore",
            "Vakkennis keuzevak",
            "punten (0-5)",
        ],
        [
            "V2 SCORE",
            "Schooldiploma keuzevak",
            "Keuzevak 2 puntenscore",
            "Vakkennis keuzevak",
            "punten (0-5)",
        ],
        [
            "V3 SCORE",
            "Schooldiploma keuzevak",
            "Keuzevak 3 puntenscore",
            "Vakkennis keuzevak",
            "punten (0-5)",
        ],
        [
            "V4 SCORE",
            "Schooldiploma keuzevak",
            "Keuzevak 4 puntenscore",
            "Vakkennis keuzevak",
            "punten (0-5)",
        ],
        [
            "V5 SCORE",
            "Schooldiploma keuzevak",
            "Keuzevak 5 puntenscore",
            "Vakkennis keuzevak",
            "punten (0-5)",
        ],
        [
            "V6 SCORE",
            "Schooldiploma keuzevak",
            "Keuzevak 6 puntenscore",
            "Vakkennis keuzevak",
            "punten (0-5)",
        ],
        [
            "V7 SCORE",
            "Schooldiploma keuzevak",
            "Keuzevak 7 puntenscore",
            "Vakkennis keuzevak",
            "punten (0-5)",
        ],
        [
            "V8 SCORE",
            "Schooldiploma keuzevak",
            "Keuzevak 8 puntenscore",
            "Vakkennis keuzevak",
            "punten (0-5)",
        ],
        [
            "V9 SCORE",
            "Schooldiploma keuzevak",
            "Keuzevak 9 puntenscore",
            "Vakkennis keuzevak",
            "punten (0-5)",
        ],
        [
            "V10 SCORE",
            "Schooldiploma keuzevak",
            "Keuzevak 10 puntenscore",
            "Vakkennis keuzevak",
            "punten (0-5)",
        ],
        [
            "COMBICIJF \nSCORE",
            "Schooldiploma",
            "Combinatiecijfer puntenscore",
            "Algemeen studieniveau",
            "punten (0-5)",
        ],
        [
            "V1-V10 + COMBICIJF (0-5)",
            "Deelscore",
            "Totaalscore vragenlijst (0-5)",
            "",
            "punten (0-5)",
        ],
        [
            "Matching Score (1-3)",
            "Matchingsvragenlijst",
            "Matchingscore (1-3)",
            "Studiemotivatie",
            "schaal 1-3",
        ],
        [
            "Vragenlijst %",
            "Deelscore",
            "Vragenlijst score percentage",
            "",
            "percentage",
        ],
        ["Matching %", "Deelscore", "Matching score percentage", "", "percentage"],
    ],
)

# NEW: Psychologie 2022-2023
make_config(
    "config_Psychologie_2022_2023.xlsx",
    [
        ("Koppel_id_kolom", "Studentnummer"),
        ("opleiding", "Psychologie"),
        ("jaar", "2022"),
        ("blad_naam", "Totaalscores met formules"),
        ("totaalscore_kolom", "Totaalscore"),
    ],
    [
        ["Toets", "Selectietoets", "Toets ruwe score", "Selectietoets", "punten"],
        [
            "Toetsscore",
            "Selectietoets",
            "Toets percentage",
            "Selectietoets",
            "percentage",
        ],
        [
            "Matching",
            "Matchingsvragenlijst",
            "Matching ruwe score",
            "Studiemotivatie",
            "punten",
        ],
        [
            "Matchingscore",
            "Matchingsvragenlijst",
            "Matching percentage",
            "Studiemotivatie",
            "percentage",
        ],
        [
            "Cijferlijst",
            "Cijferlijst",
            "Cijferlijst ruwe score",
            "Studieniveau",
            "punten",
        ],
        [
            "Cijferlijstscore",
            "Cijferlijst",
            "Cijferlijst percentage",
            "Studieniveau",
            "percentage",
        ],
    ],
)

# NEW: Psychologie 2021-2022
make_config(
    "config_Psychologie_2021_2022.xlsx",
    [
        ("Koppel_id_kolom", "Studentnummer"),
        ("opleiding", "Psychologie"),
        ("jaar", "2021"),
        ("blad_naam", "Totaalscores met formules"),
        ("totaalscore_kolom", "Totaalscore"),
    ],
    [
        ["TOETS", "Selectietoets", "Toets ruwe score", "Selectietoets", "punten"],
        [
            "Toets score",
            "Selectietoets",
            "Toets percentage",
            "Selectietoets",
            "percentage",
        ],
        [
            "MATCH",
            "Matchingsvragenlijst",
            "Matching ruwe score",
            "Studiemotivatie",
            "punten",
        ],
        [
            "Match score",
            "Matchingsvragenlijst",
            "Matching percentage",
            "Studiemotivatie",
            "percentage",
        ],
        ["CIJF", "Cijferlijst", "Cijferlijst ruwe score", "Studieniveau", "punten"],
        [
            "Cijf score",
            "Cijferlijst",
            "Cijferlijst percentage",
            "Studieniveau",
            "percentage",
        ],
    ],
)
